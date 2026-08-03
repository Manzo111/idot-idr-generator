import io
import re
import sqlite3
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from pathlib import Path
from urllib.parse import urljoin, urlparse, parse_qs, urlencode, urlunparse, quote, unquote
from io import StringIO
import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import black, white
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.colors import black, white
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.colors import black, white
BASE_DIR = Path(__file__).parent
TEMPLATE_CANDIDATES = [BASE_DIR / 'IDR_Template.xlsx', BASE_DIR / 'IDR_template.xlsx']
TEMPLATE_PATH = next((path for path in TEMPLATE_CANDIDATES if path.exists()), TEMPLATE_CANDIDATES[0])
FLAT_FLAT_PDF_TEMPLATE_CANDIDATES = [
    BASE_DIR / 'bc-628-flat-flat.pdf',
    BASE_DIR / 'BC-628-flat-flat.pdf',
    BASE_DIR / 'BC_628_flat_flat.pdf',
]
FLAT_FLAT_PDF_TEMPLATE_PATH = next(
    (path for path in FLAT_FLAT_PDF_TEMPLATE_CANDIDATES if path.exists()),
    FLAT_FLAT_PDF_TEMPLATE_CANDIDATES[0],
)

# Coordinates below are based on a standard 8.5 x 11 inch BC-628 form.
# Values are PDF points measured from the lower-left corner.
BC628_COORDS = {
    # The flattened BC-628 supplied by IDOT is landscape: 792 x 612 points.
    # PDF coordinates are measured from the lower-left corner.

    # Upper-left fields
    'date': (58, 531),
    'contractor': (58, 514),
    'weather': (58, 479),

    # Initial/date block
    'inspected_by': (353, 481),
    'inspected_date': (434, 481),
    'measured_by': (353, 464),
    'measured_date': (434, 464),
    'calculated_by': (353, 447),
    'calculated_date': (434, 447),

    # Upper-right job-information boxes
    'county': (560, 527),
    'route': (560, 505),
    'section': (681, 505),
    'district': (560, 454),
    'contract': (649, 454),
    'job': (719, 454),
    'project': (560, 420),

    # Pay-item table. This is the baseline of row 1.
    'table_top_y': 287,
    'table_row_height': 28.45,
    'item_code_x': 22,
    'fund_code_x': 107,
    'description_x': 160,
    'location_x': 290,
    'quantity_x': 420,
    'evidence_x': 512,

    # Measurement rows
    'estimated_check': (49, 118),
    'estimated_item_x': 274,
    'final_check': (49, 99),
    'final_item_x': 244,

    # Remarks box
    'remarks_x': 45,
    'remarks_y': 164,
    'remarks_width': 718,
    'remarks_height': 95,
}
BASE_URL = 'https://webapps1.dot.illinois.gov'
IDOT_HOME_URL = 'https://webapps1.dot.illinois.gov/WCTB/LBHome'
SEARCH_MAX_PAGES_PER_LETTING = 4
CACHE_TTL_SECONDS = 86400
REQUEST_TIMEOUT_SECONDS = 6
RECENT_LETTINGS_FIRST = 4
FAST_PAGES_PER_LETTING = 2
FAST_SEARCH_TIME_BUDGET_SECONDS = 18
PUBLIC_SEARCH_MAX_CANDIDATES = 2
ENABLE_PUBLIC_SEARCH_ON_FAST_LOOKUP = False
ENABLE_FULL_ARCHIVE_ON_FAST_LOOKUP = False
RESOLVE_ARCHIVE_DATES_WITH_SEARCH = False
MAX_ARCHIVE_DATES_TO_RESOLVE = 3
CONTRACT_INDEX_DB_PATH = BASE_DIR / 'idot_contract_index.sqlite'
CONTRACT_MISS_TTL_SECONDS = 3600
RECENT_INDEX_LETTINGS = 4
RECENT_INDEX_MAX_PAGES_PER_LETTING = 2
FULL_INDEX_MAX_PAGES_PER_LETTING = SEARCH_MAX_PAGES_PER_LETTING
INDEX_MAX_WORKERS = 8
RECENT_INDEX_TTL_SECONDS = 6 * 3600
FULL_INDEX_TTL_SECONDS = CACHE_TTL_SECONDS
DESCRIPTION_TEXT_FIT_RULES = [(35, 10), (70, 9), (105, 8), (140, 7), (9999, 6)]
UNIT_WORDS = ['CU YD', 'CUYD', 'SQ YD', 'SQYD', 'SQ FT', 'SQFT', 'FOOT', 'EACH', 'L SUM', 'LSUM', 'CAL DA', 'CAL MO', 'POUND', 'HOUR', 'TON', 'GALLON', 'ACRE', 'UNIT', 'SQ M', 'METER', 'LITER', 'M GAL', 'L FOOT']
EVIDENCE_BY_EXACT_ITEM_CODE = {'20101700': 'Potable source', '20200100': 'None', '21101505': 'None, topsoil taken from within R.O.W.', '25200200': 'Potable source'}
EVIDENCE_BY_SECTION = {'202': 'None', '203': 'None', '204': 'Soil from outside R.O.W.: Letter of approval from District Materials Engineer', '206': 'Approved source & shipment ticket or LIST + TICK', '207': 'Approved source & shipment ticket or LIST + TICK', '208': 'Approved source & shipment ticket or LIST + TICK', '209': 'Approved source & shipment ticket or LIST + TICK', '210': 'CERT or LA15', '213': 'None', '311': 'Approved source & shipment ticket or LIST + TICK', '312': 'HMA: DPR + TICK + TEST; CAM II: DPR + TICK + TEST; CAM/PSM: TEST', '351': 'Approved source & shipment ticket or LIST + TICK', '353': 'DPR + TICK + TEST', '354': 'DPR + TICK + TEST', '355': 'DPR + TICK + TEST', '356': 'DPR + TICK + TEST', '402': 'Approved source & shipment ticket or LIST + TICK', '407': 'DPR + TICK + TEST', '424': 'DPR + TICK + TEST', '481': 'Approved source & shipment ticket or LIST + TICK', '482': 'DPR + TICK + TEST', '501': 'None', '502': 'None', '504': 'Precast bridge slab: LIST + ILOK; Precast bridge beams: LIST + ILOK; Prestressed bridge beams: ILOK', '505': "Steel: Fabrication Inspector's Release (BBS 59) + CERT; High-strength bolts: CBM or LA15 or ILOK or TEST", '508': 'Rebar: LIST + CERT + MARK; Epoxy-coated rebar: LIST + CERT + MARK', '511': 'Concrete: DPR + TICK + TEST; Mesh: LIST + CERT', '542': 'Cast-in-place: DPR + TICK + TEST; Precast: LIST + MARK', '550': 'Concrete: LIST + MARK; Plastic: ILOK or LA15 or TEST; Clay: ILOK or LA15 or TEST', '580': 'LA15 or TEST', '606': 'DPR + TICK + TEST', '611': 'DPR + TICK + TEST', '630': 'Rail element: LIST + CERT; Steel post: CERT or LA15; End section: (LIST + CERT) or LA15; Fasteners: (MARK + CERT) or TEST; Wood post: CERT or MARK or LA15', '663': 'Dust palliative: TEST; Accelerator: CERT', '664': 'CERT or LA15', '665': 'CERT or LA15'}
EVIDENCE_RULES = [('201', ('TREE', 'REMOVAL'), 'None'), ('201', ('TEMPORARY', 'FENCE'), 'VISE'), ('201', ('FERTILIZER', 'NUTRIENT'), 'CERT (bulk) or MARK (bags)'), ('201', ('SUPPLEMENTAL', 'WATERING'), 'Potable source'), ('202', ('EARTH', 'EXCAVATION', 'WIDENING'), 'None'), ('202', ('ROCK', 'EXCAVATION'), 'None'), ('202', ('EARTH', 'EXCAVATION'), 'None'), ('203', ('ROCK', 'EXCAVATION', 'CHANNEL'), 'None'), ('203', ('CHANNEL', 'EXCAVATION'), 'None'), ('204', ('BORROW', 'EXCAVATION'), 'Soil from outside R.O.W.: Letter of approval from District Materials Engineer'), ('204', ('FURNISHED', 'EXCAVATION'), 'Soil from outside R.O.W.: Letter of approval from District Materials Engineer'), ('206', ('GRANULAR', 'EMBANKMENT', 'SPECIAL'), 'Approved source & shipment ticket or LIST + TICK'), ('207', ('POROUS', 'GRANULAR', 'EMBANKMENT'), 'Approved source & shipment ticket or LIST + TICK'), ('208', ('TRENCH', 'BACKFILL'), 'Approved source & shipment ticket or LIST + TICK'), ('209', ('POROUS', 'GRANULAR', 'BACKFILL'), 'Approved source & shipment ticket or LIST + TICK'), ('210', ('GEOTECHNICAL', 'FABRIC'), 'CERT or LA15'), ('211', ('TOPSOIL', 'FURNISH'), 'TEST'), ('211', ('COMPOST', 'FURNISH'), 'CERT'), ('213', ('EXPLORATION', 'TRENCH'), 'None'), ('250', ('SEEDING',), 'CERT or ILOK or LA15'), ('250', ('AGRICULTURAL', 'GROUND', 'LIMESTONE'), 'Approved source & shipment ticket or LIST + TICK'), ('252', ('AGRICULTURAL', 'GROUND', 'LIMESTONE'), 'Approved source & shipment ticket or LIST + TICK'), ('250', ('FERTILIZER', 'NUTRIENT'), 'CERT (bulk) or MARK (bags)'), ('252', ('FERTILIZER', 'NUTRIENT'), 'CERT (bulk) or MARK (bags)'), ('311', ('SUBBASE', 'GRANULAR', 'MATERIAL'), 'Approved source & shipment ticket or LIST + TICK'), ('312', ('STABILIZED', 'SUBBASE'), 'HMA: DPR + TICK + TEST; CAM II: DPR + TICK + TEST; CAM/PSM: TEST'), ('351', ('AGGREGATE', 'BASE', 'COURSE'), 'Approved source & shipment ticket or LIST + TICK'), ('352', ('PROCESSING', 'SOIL', 'CEMENT', 'BASE'), 'TEST'), ('352', ('CEMENT',), '(LIST or TEST) + BOL'), ('353', ('PCC', 'BASE', 'COURSE'), 'DPR + TICK + TEST'), ('354', ('PCC', 'BASE', 'COURSE', 'WIDENING'), 'DPR + TICK + TEST'), ('355', ('HMA', 'COURSE', 'WIDENING'), 'DPR + TICK + TEST'), ('356', ('HMA', 'BASE', 'COURSE', 'WIDENING'), 'DPR + TICK + TEST'), ('358', ('PREPARATION', 'BASE'), 'None'), ('358', ('AGGREGATE', 'BASE', 'REPAIR'), 'Approved source & shipment ticket or LIST + TICK'), ('402', ('AGGREGATE', 'SURFACE', 'COURSE'), 'Approved source & shipment ticket or LIST + TICK'), ('403', ('BIT', 'MATERIAL', 'PRIME', 'COAT'), '(LIST or TEST) + BOL'), ('403', ('COVER', 'COAT', 'AGGREGATE'), 'Approved source & shipment ticket or LIST + TICK'), ('406', ('AGGREGATE', 'PRIME', 'COAT'), 'Approved source & shipment ticket or LIST + TICK'), ('406', ('BIT', 'MATERIAL', 'PRIME', 'COAT'), '(LIST or TEST) + BOL'), ('408', ('POLYMERIZED', 'PRIME', 'COAT'), '(LIST or TEST) + BOL'), ('406', ('MIX', 'CRACK', 'JOINT'), 'DPR + TICK + TEST'), ('406', ('LEVELING', 'BINDER'), 'DPR + TICK + TEST'), ('406', ('HMA', 'BINDER', 'COURSE'), 'DPR + TICK + TEST'), ('406', ('HMA', 'SURFACE', 'COURSE'), 'DPR + TICK + TEST'), ('407', ('HMA', 'PAVEMENT', 'FULL', 'DEPTH'), 'DPR + TICK + TEST'), ('408', ('INCIDENTAL', 'HMA', 'SURFACING'), 'DPR + TICK + TEST'), ('420', ('WELDED', 'WIRE', 'REINFORCEMENT'), 'LIST + CERT'), ('420', ('PROTECTIVE', 'COAT'), 'LA15 or ILOK or TEST or CBM'), ('420', ('BRIDGE', 'APPROACH', 'PAVEMENT'), 'DPR + TICK + TEST'), ('420', ('PCC', 'PAVEMENT'), 'DPR + TICK + TEST'), ('421', ('WIDE', 'FLANGE', 'BEAM', 'TERMINAL'), 'Concrete: DPR + TICK + TEST; Rebar: LIST + CERT + MARK; Epoxy-coated rebar: LIST + CERT + MARK; Steel beam: BBS 59 + CERT'), ('421', ('PAVEMENT', 'REINFORCEMENT'), 'LIST + CERT + MARK'), ('421', ('PROTECTIVE', 'COAT'), 'LA15 or ILOK or TEST or CBM'), ('421', ('CONTINUOUSLY', 'REINFORCED', 'PCC'), 'DPR + TICK + TEST'), ('424', ('PCC', 'SIDEWALK'), 'DPR + TICK + TEST'), ('481', ('AGGREGATE', 'SHOULDER'), 'Approved source & shipment ticket or LIST + TICK'), ('482', ('HMA', 'SHOULDER'), 'DPR + TICK + TEST'), ('501', ('CONCRETE', 'REMOVAL'), 'None'), ('502', ('STRUCTURE', 'EXCAVATION'), 'None'), ('503', ('RUBBED', 'FINISH'), 'None'), ('503', ('CLASS', 'MS', 'CONCRETE'), 'DPR + TICK + TEST'), ('503', ('CONCRETE', 'STRUCTURE'), 'DPR + TICK + TEST'), ('503', ('CONCRETE', 'SUPERSTRUCTURE'), 'DPR + TICK + TEST'), ('504', ('PRECAST', 'CONCRETE', 'BRIDGE', 'SLAB'), 'Precast bridge slab: LIST + ILOK; Precast bridge beams: LIST + ILOK; Prestressed bridge beams: ILOK'), ('505', ('STRUCTURAL', 'STEEL'), "Steel: Fabrication Inspector's Release (BBS 59) + CERT; High-strength bolts: CBM or LA15 or ILOK or TEST"), ('508', ('REINFORCEMENT', 'BAR'), 'Rebar: LIST + CERT + MARK; Epoxy-coated rebar: LIST + CERT + MARK'), ('509', ('STEEL', 'RAILING'), 'Steel railing: CBM; Fasteners: CBM or LA15 or ILOK or TEST; Posts/anchors: CERT or LA15'), ('509', ('ALUMINUM', 'RAILING'), 'Aluminum railing: CERT or LA15; Fasteners: CBM or LA15 or ILOK or TEST; Posts/anchors: CERT or LA15'), ('509', ('PEDESTRIAN', 'RAILING'), 'CERT or LA15'), ('509', ('BICYCLE', 'RAILING'), 'CERT or LA15'), ('511', ('SLOPE', 'WALL'), 'Concrete: DPR + TICK + TEST; Mesh: LIST + CERT'), ('512', ('FURNISHING', 'PILE'), 'Precast concrete: LIST + ILOK; Prestressed concrete: ILOK; Steel H/metal shell: CERT or LA15 or ILOK; Timber: CERT or MARK or LA15'), ('512', ('DRIVE', 'PILE'), 'None'), ('542', ('CONCRETE', 'COLLAR'), 'Cast-in-place: DPR + TICK + TEST; Precast: LIST + MARK'), ('550', ('STORM', 'SEWER'), 'Concrete: LIST + MARK; Plastic: ILOK or LA15 or TEST; Clay: ILOK or LA15 or TEST'), ('580', ('MEMBRANE', 'WATERPROOFING'), 'LA15 or TEST'), ('606', ('CONCRETE', 'CURB'), 'DPR + TICK + TEST'), ('606', ('CONCRETE', 'GUTTER'), 'DPR + TICK + TEST'), ('606', ('PAVED', 'DITCH'), 'DPR + TICK + TEST'), ('611', ('CLASS', 'SI', 'CONCRETE'), 'DPR + TICK + TEST'), ('630', ('STEEL', 'PLATE', 'BEAM', 'GUARD', 'RAIL'), 'Rail element: LIST + CERT; Steel post: CERT or LA15; End section: (LIST + CERT) or LA15; Fasteners: (MARK + CERT) or TEST; Wood post: CERT or MARK or LA15'), ('663', ('CALCIUM', 'CHLORIDE'), 'Dust palliative: TEST; Accelerator: CERT'), ('664', ('CHAIN', 'LINK', 'FENCE'), 'CERT or LA15'), ('665', ('WOVEN', 'WIRE', 'FENCE'), 'CERT or LA15'), ('780', ('THERMOPLASTIC', 'PAVEMENT', 'MARKING'), 'LA15 or ILOK or CBM'), ('780', ('PREFORMED', 'PLASTIC', 'PAVEMENT', 'MARKING'), 'CERT or LA15 or ILOK or CBM'), ('780', ('EPOXY', 'PAVEMENT', 'MARKING'), 'LA15 or CBM'), ('780', ('MODIFIED', 'URETHANE', 'MARKING'), 'LA15 or CBM')]

def normalize_evidence_match_text(value):
    value = clean_line(value).upper()
    value = value.replace('&', ' AND ')
    value = re.sub('[^A-Z0-9]+', ' ', value)
    return re.sub('\\s+', ' ', value).strip()

def get_item_section(item_code):
    digits = re.sub('\\D', '', clean_line(item_code))
    return digits[:3] if len(digits) >= 3 else ''

def get_evidence_of_material_inspection(item_code, item_description):
    code = normalize_pay_item_code(item_code)
    if code in EVIDENCE_BY_EXACT_ITEM_CODE:
        return EVIDENCE_BY_EXACT_ITEM_CODE[code]
    section = get_item_section(code)
    if not section:
        return ''
    if section in EVIDENCE_BY_SECTION:
        return EVIDENCE_BY_SECTION[section]
    description = normalize_evidence_match_text(item_description)
    if not description:
        return ''
    for rule_section, phrases, evidence in EVIDENCE_RULES:
        if section != rule_section:
            continue
        if all((normalize_evidence_match_text(phrase) in description for phrase in phrases)):
            return evidence
    return ''

def get_headers():
    return {'User-Agent': 'Mozilla/5.0 IDR Generator', 'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'}

def make_session():
    session = requests.Session()
    session.headers.update(get_headers())
    return session

def make_soup(markup):
    """Use lxml when installed, but keep working if Streamlit lacks it."""
    try:
        return BeautifulSoup(markup, 'lxml')
    except Exception:
        return BeautifulSoup(markup, 'html.parser')

def get_html(session, url):
    response = session.get(url, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    return response.text

def get_response(session, url):
    response = session.get(url, timeout=REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    return response

def absolute_url(href, base=BASE_URL + '/WCTB/'):
    return urljoin(base, href)

def set_query_param(url, key, value):
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    query[key] = [str(value)]
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, urlencode(query, doseq=True), parsed.fragment))

def clean_line(text):
    text = str(text)
    if text.lower() in ['nan', 'none']:
        return ''
    text = text.replace('\xa0', ' ')
    text = re.sub('\\s+', ' ', text)
    return text.strip()

def strip_bullet(text):
    text = clean_line(text)
    text = text.lstrip('*').strip()
    return text

def decode_content(content):
    for encoding in ['utf-8', 'utf-16', 'cp1252', 'latin1']:
        try:
            return content.decode(encoding)
        except Exception:
            continue
    return content.decode('latin1', errors='ignore')

def parse_region_project_from_values(values, metadata):
    cleaned_values = [clean_line(x) for x in values if clean_line(x)]
    if len(cleaned_values) >= 1:
        metadata['region'] = cleaned_values[0]
    if len(cleaned_values) >= 2:
        metadata['district'] = cleaned_values[1]
    if len(cleaned_values) >= 3:
        metadata['dbe_percent'] = cleaned_values[2]
    if len(cleaned_values) >= 4:
        metadata['vbp_percent'] = cleaned_values[3]
    if len(cleaned_values) >= 5:
        metadata['federal_project'] = cleaned_values[4]

def parse_region_project_from_lines(lines, metadata):
    for i, line in enumerate(lines):
        line_lower = line.lower()
        if 'region' in line_lower and 'district' in line_lower and ('federal project' in line_lower):
            if i + 1 < len(lines):
                values = lines[i + 1].split()
                parse_region_project_from_values(values, metadata)
            return

def parse_region_project_from_flat_text(flat_text, metadata):
    pattern = re.compile('Region\\s+District\\s+DBE\\s*%\\s+VBP\\s*%\\s+Federal\\s+Project\\s*#\\s+(?P<region>\\S+)\\s+(?P<district>\\S+)\\s+(?P<dbe>\\S+)\\s+(?P<vbp>\\S+)\\s+(?P<federal_project>\\S+)', re.IGNORECASE)
    match = pattern.search(flat_text)
    if not match:
        return
    metadata['region'] = clean_line(match.group('region'))
    metadata['district'] = clean_line(match.group('district'))
    metadata['dbe_percent'] = clean_line(match.group('dbe'))
    metadata['vbp_percent'] = clean_line(match.group('vbp'))
    metadata['federal_project'] = clean_line(match.group('federal_project'))

def parse_county_route_from_values(values, metadata):
    cleaned_values = []
    for value in values:
        value = strip_bullet(value)
        if not value:
            continue
        value_lower = value.lower()
        if 'county' in value_lower and 'key route' in value_lower:
            continue
        if value_lower.startswith('contract specifics'):
            continue
        cleaned_values.append(value)
    if len(cleaned_values) >= 1:
        metadata['county'] = cleaned_values[0]
    if len(cleaned_values) >= 2:
        metadata['key_route'] = cleaned_values[1]
    if len(cleaned_values) >= 3:
        metadata['marked_route'] = cleaned_values[2]
    if len(cleaned_values) >= 4:
        metadata['website_section'] = cleaned_values[3]
    if len(cleaned_values) >= 5:
        metadata['state_job'] = cleaned_values[4]
    if len(cleaned_values) >= 6:
        metadata['pps'] = ' / '.join(cleaned_values[5:])

def parse_county_route_from_lines(lines, metadata):
    """
    Most reliable parser for this IDOT page:
    BeautifulSoup.stripped_strings usually returns this section as separate lines:

    County(s) Key Route(s) Marked Route(s) Section(s) State Job #(s) PPS #(s)
    Cook
    FAP 350
    IL 50 (CICERO AVE)
    FAP 0350 22 RS
    C-91-308-22
    1-80995-0000

    This handles sections that are NOT shaped like 2019-161-W.
    Example: FAP 0350 22 RS
    """
    for i, line in enumerate(lines):
        line_lower = line.lower()
        if 'county' in line_lower and 'key route' in line_lower and ('marked route' in line_lower) and ('state job' in line_lower):
            values = []
            for next_line in lines[i + 1:]:
                next_line = strip_bullet(next_line)
                if not next_line:
                    continue
                if next_line.lower().startswith('contract specifics'):
                    break
                values.append(next_line)
            parse_county_route_from_values(values, metadata)
            return

def parse_county_route_from_flat_text(flat_text, metadata):
    pattern = re.compile('County\\s*\\(s\\)\\s+Key\\s+Route\\s*\\(s\\)\\s+Marked\\s+Route\\s*\\(s\\)\\s+Section\\s*\\(s\\)\\s+State\\s+Job\\s*#\\s*\\(s\\)\\s+PPS\\s*#\\s*\\(s\\)\\s+(?P<block>.*?)\\s+Contract\\s+Specifics', re.IGNORECASE | re.DOTALL)
    match = pattern.search(flat_text)
    if not match:
        return
    block = clean_line(match.group('block'))
    bullet_values = re.findall('\\*\\s*([^*]+?)(?=\\s*\\*|$)', block)
    if bullet_values:
        parse_county_route_from_values(bullet_values, metadata)
        return
    state_job_match = re.search('\\bC-\\d{2}-\\d{3}-\\d{2}\\b', block)
    if not state_job_match:
        return
    before_state_job = block[:state_job_match.start()].strip()
    after_state_job = block[state_job_match.end():].strip()
    metadata['state_job'] = state_job_match.group(0)
    key_route_match = re.search('\\b(?:FAP|FAU|FAS|FAI|SBI|CH|TR|IL|US|I)\\s*[A-Z0-9.-]+\\b', before_state_job, re.IGNORECASE)
    if not key_route_match:
        return
    metadata['county'] = clean_line(before_state_job[:key_route_match.start()])
    metadata['key_route'] = clean_line(key_route_match.group(0))
    after_key_route = before_state_job[key_route_match.end():].strip()
    section_patterns = ['\\b\\d{4}-[A-Z0-9-]+\\b', '\\b(?:FAP|FAU|FAS|FAI|SBI)\\s+\\d{3,4}\\s+[A-Z0-9]+\\s+[A-Z0-9]+\\b', '\\b\\d{2,4}\\s+[A-Z0-9-]+\\b']
    best_section_match = None
    for pattern_text in section_patterns:
        matches = list(re.finditer(pattern_text, after_key_route, re.IGNORECASE))
        if matches:
            best_section_match = matches[-1]
            break
    if best_section_match:
        metadata['marked_route'] = clean_line(after_key_route[:best_section_match.start()])
        metadata['website_section'] = clean_line(best_section_match.group(0))
    else:
        metadata['marked_route'] = clean_line(after_key_route)
    if after_state_job:
        metadata['pps'] = clean_line(after_state_job)

def parse_metadata_from_contract_page(html, contract_url):
    soup = make_soup(html)
    lines = [clean_line(x) for x in soup.stripped_strings if clean_line(x)]
    flat_text = clean_line(soup.get_text(' '))
    metadata = {'contract_url': contract_url, 'item_contract': '', 'letting_date': '', 'region': '', 'district': '', 'dbe_percent': '', 'vbp_percent': '', 'federal_project': '', 'county': '', 'key_route': '', 'marked_route': '', 'website_section': '', 'state_job': '', 'pps': '', 'working_days': ''}
    letting_match = re.search('\\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\\s+\\d{1,2},\\s+\\d{4}\\s+Letting\\s+\\d{1,2}:\\d{2}\\s+(?:AM|PM)', flat_text, re.IGNORECASE)
    if letting_match:
        metadata['letting_date'] = clean_line(letting_match.group(0))
    contract_match = re.search('\\b\\d{3}-[A-Z0-9]{5}\\b', flat_text)
    if contract_match:
        metadata['item_contract'] = clean_line(contract_match.group(0))
    parse_region_project_from_lines(lines, metadata)
    parse_county_route_from_lines(lines, metadata)
    if not metadata['region'] or not metadata['district'] or (not metadata['federal_project']):
        parse_region_project_from_flat_text(flat_text, metadata)
    if not metadata['county'] or not metadata['key_route'] or (not metadata['marked_route']) or (not metadata['website_section']) or (not metadata['state_job']):
        parse_county_route_from_flat_text(flat_text, metadata)
    working_days_match = re.search('(\\d+)\\s+Working\\s+Days', flat_text, re.IGNORECASE)
    if working_days_match:
        metadata['working_days'] = working_days_match.group(1)
    return metadata

def normalize_contract_input(value):
    value = clean_line(value).upper()
    value = value.replace(' ', '')
    return value

def extract_contract_label(text):
    text = clean_line(text).upper()
    match = re.search('\\b\\d{3}-[A-Z0-9]{5}\\b', text)
    if match:
        return match.group(0)
    return ''

def contract_matches(user_job_number, contract_label):
    user_job_number = normalize_contract_input(user_job_number)
    contract_label = normalize_contract_input(contract_label)
    if not user_job_number or not contract_label:
        return False
    if user_job_number == contract_label:
        return True
    contract_suffix = contract_label.split('-')[-1]
    if user_job_number == contract_suffix:
        return True
    return False

def get_page_signature(contract_links):
    labels = []
    for item in contract_links[:15]:
        labels.append(item.get('label', '') + '|' + item.get('url', ''))
    return '||'.join(labels)

def extract_archive_dates_from_home(html):
    soup = make_soup(html)
    text = clean_line(soup.get_text(' '))
    start = text.find('Transportation Bulletin Archives')
    if start != -1:
        text = text[start:]
    end = text.find('Prior Lettings')
    if end != -1:
        text = text[:end]
    date_pattern = re.compile('\\b(January|February|March|April|May|June|July|August|September|October|November|December)\\s+\\d{1,2},\\s+\\d{4}\\b')
    dates = []
    for match in date_pattern.finditer(text):
        date_text = match.group(0)
        if date_text not in dates:
            dates.append(date_text)
    return dates

def extract_letting_links_from_any_html(html):
    links = []
    raw_matches = re.findall('https?://(?:webapps1|webapps)\\.dot\\.illinois\\.gov/WCTB/LbLettingDetail/Index/[0-9a-fA-F-]+', html)
    for url in raw_matches:
        links.append({'text': 'Letting', 'url': url.replace('webapps.dot.illinois.gov', 'webapps1.dot.illinois.gov'), 'source': 'raw-html'})
    relative_matches = re.findall('/WCTB/LbLettingDetail/Index/[0-9a-fA-F-]+', html)
    for href in relative_matches:
        links.append({'text': 'Letting', 'url': absolute_url(href), 'source': 'raw-html-relative'})
    soup = make_soup(html)
    for tag in soup.find_all(True):
        visible_text = clean_line(tag.get_text(' '))
        for attr_name, attr_value in tag.attrs.items():
            if isinstance(attr_value, list):
                attr_value = ' '.join(attr_value)
            attr_value = str(attr_value)
            if 'LbLettingDetail/Index' in attr_value or 'lblettingdetail/index' in attr_value.lower():
                matches = re.findall('(?:https?://(?:webapps1|webapps)\\.dot\\.illinois\\.gov)?/?WCTB/LbLettingDetail/Index/[0-9a-fA-F-]+', attr_value)
                for href in matches:
                    if href.startswith('http'):
                        url = href
                    elif href.startswith('/'):
                        url = BASE_URL + href
                    else:
                        url = BASE_URL + '/' + href
                    url = url.replace('webapps.dot.illinois.gov', 'webapps1.dot.illinois.gov')
                    links.append({'text': visible_text or 'Letting', 'url': url, 'source': f'tag-attr-{attr_name}'})
    unique_links = []
    seen = set()
    for item in links:
        url = item['url']
        if url in seen:
            continue
        seen.add(url)
        unique_links.append(item)
    return unique_links

def extract_contract_urls_from_search_text(text):
    urls = []
    urls.extend(re.findall('https://(?:webapps1|webapps)\\.dot\\.illinois\\.gov/WCTB/LbContractDetail/Index/[^\\"\'<> \\n\\r]+', text, flags=re.IGNORECASE))
    urls.extend(re.findall('<link>(https://(?:webapps1|webapps)\\.dot\\.illinois\\.gov/WCTB/LbContractDetail/Index/.*?)</link>', text, flags=re.IGNORECASE))
    cleaned_urls = []
    for url in urls:
        url = url.replace('&amp;', '&')
        url = url.replace('webapps.dot.illinois.gov', 'webapps1.dot.illinois.gov')
        url = url.split('&form=')[0]
        url = url.split('&ved=')[0]
        url = url.strip()
        if 'LbContractDetail/Index' not in url:
            continue
        if url not in cleaned_urls:
            cleaned_urls.append(url)
    return cleaned_urls

def extract_letting_urls_from_search_text(text):
    urls = []
    urls.extend(re.findall('https://(?:webapps1|webapps)\\.dot\\.illinois\\.gov/WCTB/LbLettingDetail/Index/[0-9a-fA-F-]+', text, flags=re.IGNORECASE))
    urls.extend(re.findall('<link>(https://(?:webapps1|webapps)\\.dot\\.illinois\\.gov/WCTB/LbLettingDetail/Index/[0-9a-fA-F-]+)</link>', text, flags=re.IGNORECASE))
    cleaned_urls = []
    for url in urls:
        url = url.replace('&amp;', '&')
        url = url.replace('webapps.dot.illinois.gov', 'webapps1.dot.illinois.gov')
        url = url.split('&form=')[0]
        url = url.split('&ved=')[0]
        url = url.strip()
        if 'LbLettingDetail/Index' not in url:
            continue
        if url not in cleaned_urls:
            cleaned_urls.append(url)
    return cleaned_urls

def bing_rss_search(session, query):
    search_url = 'https://www.bing.com/search?q=' + quote(query) + '&format=rss'
    response = session.get(search_url, timeout=REQUEST_TIMEOUT_SECONDS, headers={'User-Agent': 'Mozilla/5.0 IDR Generator', 'Accept': 'application/rss+xml,application/xml,text/xml,*/*'})
    response.raise_for_status()
    return (response.text, search_url)

def resolve_archive_date_to_letting_url(session, date_text):
    queries = [f'site:webapps1.dot.illinois.gov/WCTB/LbLettingDetail/Index "{date_text} Letting"', f'site:webapps.dot.illinois.gov/WCTB/LbLettingDetail/Index "{date_text} Letting"', f'"{date_text} Letting 12:00 PM" "LbLettingDetail"']
    for query in queries:
        try:
            text, _ = bing_rss_search(session, query)
            letting_urls = extract_letting_urls_from_search_text(text)
            for url in letting_urls:
                try:
                    html = get_html(session, url)
                    page_text = clean_line(make_soup(html).get_text(' '))
                    if date_text.lower() in page_text.lower():
                        return url
                except Exception:
                    continue
        except Exception:
            continue
    return ''

def get_current_letting_link(session, html):
    links = extract_letting_links_from_any_html(html)
    if links:
        links[0]['text'] = 'Current Notice of Letting'
        links[0]['source'] = 'current'
        return links[0]
    return None

def get_all_archive_letting_links_newest_first(session):
    """
    Fast version of the current/archive letting list.

    The old version tried to resolve every archive date through Bing when a
    direct link was not obvious. That made a cold search slow because it could
    trigger many search-engine requests before the user job number was even
    checked.

    This version uses direct IDOT links first. If you absolutely need old
    archive dates that are not directly linked, set
    RESOLVE_ARCHIVE_DATES_WITH_SEARCH = True near the top of the file.
    """
    home_html = get_html(session, IDOT_HOME_URL)
    direct_links = extract_letting_links_from_any_html(home_html)
    current = get_current_letting_link(session, home_html)
    final_links = []
    seen_urls = set()
    if current and current['url'] not in seen_urls:
        final_links.append(current)
        seen_urls.add(current['url'])
    for link in direct_links:
        if link['url'] not in seen_urls:
            final_links.append(link)
            seen_urls.add(link['url'])
    if RESOLVE_ARCHIVE_DATES_WITH_SEARCH:
        archive_dates = extract_archive_dates_from_home(home_html)
        resolved_count = 0
        for date_text in archive_dates:
            if resolved_count >= MAX_ARCHIVE_DATES_TO_RESOLVE:
                break
            found_url = ''
            for link in direct_links:
                if date_text.lower() in link.get('text', '').lower():
                    found_url = link['url']
                    break
            if not found_url:
                found_url = resolve_archive_date_to_letting_url(session, date_text)
                resolved_count += 1
            if found_url and found_url not in seen_urls:
                seen_urls.add(found_url)
                final_links.append({'text': date_text, 'url': found_url, 'source': 'archive-date-search-resolved'})
    return final_links

def get_contract_links_from_letting_page(html):
    soup = make_soup(html)
    contract_links = []
    for a in soup.find_all('a', href=True):
        href = a['href']
        if 'lbcontractdetail' not in href.lower():
            continue
        link_text = clean_line(a.get_text(' '))
        label = extract_contract_label(link_text)
        parent_row = a.find_parent('tr')
        row_text = ''
        if parent_row is not None:
            row_text = clean_line(parent_row.get_text(' '))
            if not label:
                label = extract_contract_label(row_text)
        contract_links.append({'label': label, 'url': absolute_url(href), 'text': link_text, 'row_text': row_text})
    return contract_links

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def cached_archive_letting_links_newest_first():
    """
    Cache the current/archive letting list so the app does not re-download and
    re-resolve archive dates on every job search.
    """
    cached_session = make_session()
    return get_all_archive_letting_links_newest_first(cached_session)

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def cached_contract_links_for_page(page_url):
    """
    Cache one letting result page. This is the biggest repeat-work saver because
    each letting can have multiple pages, and users often search several jobs
    from the same letting.
    """
    cached_session = make_session()
    html = get_html(cached_session, page_url)
    return get_contract_links_from_letting_page(html)

def make_contract_result(contract, letting, page_num, source_suffix=''):
    source = letting.get('source', '')
    if source_suffix:
        source = f'{source}-{source_suffix}' if source else source_suffix
    return {'label': contract.get('label', ''), 'url': contract.get('url', ''), 'letting': letting.get('text', ''), 'letting_url': letting.get('url', ''), 'page': page_num, 'source': source}

def get_contract_db_connection():
    """Create the local SQLite index database if needed and return a connection."""
    conn = sqlite3.connect(CONTRACT_INDEX_DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute('PRAGMA journal_mode=WAL')
        conn.execute('PRAGMA synchronous=NORMAL')
    except Exception:
        pass
    conn.execute('\n        CREATE TABLE IF NOT EXISTS contracts (\n            contract_key TEXT PRIMARY KEY,\n            label TEXT,\n            url TEXT NOT NULL,\n            letting TEXT,\n            letting_url TEXT,\n            page TEXT,\n            source TEXT,\n            updated_at REAL\n        )\n        ')
    conn.execute('\n        CREATE TABLE IF NOT EXISTS contract_misses (\n            query TEXT PRIMARY KEY,\n            updated_at REAL\n        )\n        ')
    conn.execute('\n        CREATE TABLE IF NOT EXISTS index_meta (\n            key TEXT PRIMARY KEY,\n            value TEXT,\n            updated_at REAL\n        )\n        ')
    conn.commit()
    return conn

def sqlite_lookup_contract(user_job_number):
    query = normalize_contract_input(user_job_number)
    if not query:
        return None
    try:
        with get_contract_db_connection() as conn:
            row = conn.execute('\n                SELECT label, url, letting, letting_url, page, source\n                FROM contracts\n                WHERE contract_key = ?\n                LIMIT 1\n                ', (query,)).fetchone()
        if row is None:
            return None
        return {'label': clean_line(row['label']), 'url': clean_line(row['url']), 'letting': clean_line(row['letting']), 'letting_url': clean_line(row['letting_url']), 'page': clean_line(row['page']), 'source': clean_line(row['source'] or 'sqlite-index')}
    except Exception:
        return None

def sqlite_save_contract_result(label, result, source_suffix=''):
    normalized_label = normalize_contract_input(label or result.get('label', ''))
    url = clean_line(result.get('url', ''))
    if not normalized_label or not url:
        return
    source = clean_line(result.get('source', ''))
    if source_suffix:
        source = f'{source}-{source_suffix}' if source else source_suffix
    keys = [normalized_label]
    if '-' in normalized_label:
        suffix = normalized_label.split('-')[-1]
        if suffix and suffix not in keys:
            keys.append(suffix)
    now = time.time()
    try:
        with get_contract_db_connection() as conn:
            for contract_key in keys:
                conn.execute('\n                    INSERT OR REPLACE INTO contracts\n                    (contract_key, label, url, letting, letting_url, page, source, updated_at)\n                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)\n                    ', (contract_key, normalized_label, url, clean_line(result.get('letting', '')), clean_line(result.get('letting_url', '')), str(result.get('page', '')), source, now))
            conn.execute('DELETE FROM contract_misses WHERE query IN ({})'.format(','.join(['?'] * len(keys))), keys)
            conn.commit()
    except Exception:
        pass

def sqlite_save_contract_links(letting, page_num, contract_links, source_suffix='sqlite-index'):
    saved_count = 0
    for contract in contract_links:
        label = contract.get('label', '')
        if not normalize_contract_input(label):
            continue
        result = make_contract_result(contract, letting, page_num, source_suffix=source_suffix)
        sqlite_save_contract_result(label, result)
        saved_count += 1
    return saved_count

def sqlite_is_recent_miss(user_job_number):
    query = normalize_contract_input(user_job_number)
    if not query:
        return False
    try:
        with get_contract_db_connection() as conn:
            row = conn.execute('SELECT updated_at FROM contract_misses WHERE query = ? LIMIT 1', (query,)).fetchone()
        if row is None:
            return False
        return time.time() - float(row['updated_at']) < CONTRACT_MISS_TTL_SECONDS
    except Exception:
        return False

def sqlite_record_miss(user_job_number):
    query = normalize_contract_input(user_job_number)
    if not query:
        return
    try:
        with get_contract_db_connection() as conn:
            conn.execute('\n                INSERT OR REPLACE INTO contract_misses (query, updated_at)\n                VALUES (?, ?)\n                ', (query, time.time()))
            conn.commit()
    except Exception:
        pass

def sqlite_clear_contract_misses():
    """Clear remembered failed searches after rebuilding the contract index."""
    try:
        with get_contract_db_connection() as conn:
            conn.execute('DELETE FROM contract_misses')
            conn.commit()
    except Exception:
        pass

def sqlite_get_meta(key):
    try:
        with get_contract_db_connection() as conn:
            row = conn.execute('SELECT value, updated_at FROM index_meta WHERE key = ? LIMIT 1', (key,)).fetchone()
        if row is None:
            return ('', 0.0)
        return (clean_line(row['value']), float(row['updated_at'] or 0))
    except Exception:
        return ('', 0.0)

def sqlite_set_meta(key, value):
    try:
        with get_contract_db_connection() as conn:
            conn.execute('\n                INSERT OR REPLACE INTO index_meta (key, value, updated_at)\n                VALUES (?, ?, ?)\n                ', (key, str(value), time.time()))
            conn.commit()
    except Exception:
        pass

def make_letting_signature(letting_links):
    pieces = []
    for letting in letting_links:
        pieces.append(clean_line(letting.get('text', '')) + '|' + clean_line(letting.get('url', '')))
    return '||'.join(pieces)

def fetch_contract_links_for_index(page_url):
    """Thread worker: fetch and parse one IDOT letting page."""
    try:
        session = make_session()
        html = get_html(session, page_url)
        return get_contract_links_from_letting_page(html)
    except Exception:
        return []

def build_sqlite_contract_index_for_lettings(letting_links, max_pages_per_letting, meta_key='', meta_ttl_seconds=0, force=False, source_suffix='sqlite-index'):
    """
    Build/update the persistent SQLite index.

    This keeps the app UI unchanged: the user still only enters the job number.
    The indexing is automatic and stored in idot_contract_index.sqlite.
    """
    if not letting_links:
        return {'checked_lettings': 0, 'checked_pages': 0, 'saved_contracts': 0}
    signature = make_letting_signature(letting_links)
    if meta_key and (not force):
        saved_signature, updated_at = sqlite_get_meta(meta_key)
        if saved_signature == signature and time.time() - updated_at < meta_ttl_seconds:
            return {'checked_lettings': 0, 'checked_pages': 0, 'saved_contracts': 0, 'skipped': True}
    page_jobs = []
    for letting in letting_links:
        letting_url = letting.get('url', '')
        if not letting_url:
            continue
        for page_num in range(1, max_pages_per_letting + 1):
            page_url = letting_url if page_num == 1 else set_query_param(letting_url, 'page', page_num)
            page_jobs.append((letting, page_num, page_url))
    page_results = []
    workers = max(1, min(INDEX_MAX_WORKERS, len(page_jobs)))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_map = {executor.submit(fetch_contract_links_for_index, page_url): (letting, page_num, page_url) for letting, page_num, page_url in page_jobs}
        for future in as_completed(future_map):
            letting, page_num, page_url = future_map[future]
            try:
                contract_links = future.result()
            except Exception:
                contract_links = []
            page_results.append((letting, page_num, page_url, contract_links))
    grouped = {}
    letting_order = []
    for letting, page_num, page_url, contract_links in page_results:
        letting_url = letting.get('url', '')
        if letting_url not in grouped:
            grouped[letting_url] = []
            letting_order.append(letting_url)
        grouped[letting_url].append((page_num, letting, page_url, contract_links))
    checked_lettings = 0
    checked_pages = 0
    saved_contracts = 0
    for letting_url in letting_order:
        checked_lettings += 1
        seen_page_signatures = set()
        for page_num, letting, page_url, contract_links in sorted(grouped[letting_url], key=lambda x: x[0]):
            if not contract_links:
                break
            page_signature = get_page_signature(contract_links)
            if page_signature in seen_page_signatures:
                break
            seen_page_signatures.add(page_signature)
            checked_pages += 1
            saved_contracts += sqlite_save_contract_links(letting=letting, page_num=page_num, contract_links=contract_links, source_suffix=source_suffix)
    if meta_key:
        sqlite_set_meta(meta_key, signature)
    return {'checked_lettings': checked_lettings, 'checked_pages': checked_pages, 'saved_contracts': saved_contracts}

def ensure_recent_sqlite_contract_index(letting_links):
    recent_lettings = letting_links[:RECENT_INDEX_LETTINGS]
    return build_sqlite_contract_index_for_lettings(recent_lettings, max_pages_per_letting=RECENT_INDEX_MAX_PAGES_PER_LETTING, meta_key='recent_contract_index', meta_ttl_seconds=RECENT_INDEX_TTL_SECONDS, force=False, source_suffix='sqlite-recent-index')

def ensure_full_sqlite_contract_index(letting_links):
    return build_sqlite_contract_index_for_lettings(letting_links, max_pages_per_letting=FULL_INDEX_MAX_PAGES_PER_LETTING, meta_key='full_contract_index', meta_ttl_seconds=FULL_INDEX_TTL_SECONDS, force=False, source_suffix='sqlite-full-index')

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def cached_contract_lookup_index():
    """
    Read the SQLite index into memory without building/refreshing it.
    Building the full archive index inside lookup is intentionally disabled.
    """
    contract_index = {}
    checked_lettings = 0
    checked_pages = 0
    try:
        with get_contract_db_connection() as conn:
            rows = conn.execute('\n                SELECT contract_key, label, url, letting, letting_url, page, source\n                FROM contracts\n                ').fetchall()
        for row in rows:
            contract_index[row['contract_key']] = {'label': clean_line(row['label']), 'url': clean_line(row['url']), 'letting': clean_line(row['letting']), 'letting_url': clean_line(row['letting_url']), 'page': clean_line(row['page']), 'source': clean_line(row['source'] or 'sqlite-index')}
    except Exception:
        pass
    return {'contracts': contract_index, 'checked_lettings': checked_lettings, 'checked_pages': checked_pages}

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def cached_public_contract_search(user_job_number):
    cached_session = make_session()
    return find_contract_detail_url_from_public_search(cached_session, user_job_number)

def search_time_exceeded(start_time):
    if start_time is None:
        return False
    return time.time() - start_time >= FAST_SEARCH_TIME_BUDGET_SECONDS

def letting_page_matches_contract(session, letting, user_job_number, max_pages=None, start_time=None):
    """
    Fast direct IDOT page scan.

    Important: this does NOT walk every archive page by default. It scans only a
    small number of pages for each recent letting, saves anything it sees into
    SQLite, and returns as soon as it finds the requested contract.
    """
    seen_page_signatures = set()
    max_pages = max_pages or FAST_PAGES_PER_LETTING
    for page_num in range(1, max_pages + 1):
        if search_time_exceeded(start_time):
            break
        if page_num == 1:
            page_url = letting['url']
        else:
            page_url = set_query_param(letting['url'], 'page', page_num)
        try:
            contract_links = cached_contract_links_for_page(page_url)
        except Exception:
            try:
                html = get_html(session, page_url)
                contract_links = get_contract_links_from_letting_page(html)
            except Exception:
                break
        if not contract_links:
            break
        page_signature = get_page_signature(contract_links)
        if page_signature in seen_page_signatures:
            break
        seen_page_signatures.add(page_signature)
        sqlite_save_contract_links(letting=letting, page_num=page_num, contract_links=contract_links, source_suffix='fast-visited-page')
        for contract in contract_links:
            label = contract.get('label', '')
            if contract_matches(user_job_number, label):
                result = make_contract_result(contract, letting, page_num, source_suffix='fast-page')
                sqlite_save_contract_result(label, result)
                return result
    return None

def find_contract_detail_url_from_public_search(session, user_job_number):
    """
    Very small public-search fallback.

    This is disabled by default through ENABLE_PUBLIC_SEARCH_ON_FAST_LOOKUP
    because search engines were one of the biggest reasons the lookup felt slow.
    """
    user_job_number = normalize_contract_input(user_job_number)
    search_query = f'"{user_job_number}" "LbContractDetail/Index" "webapps1.dot.illinois.gov/WCTB"'
    candidate_urls = []
    try:
        text, _ = bing_rss_search(session, search_query)
        candidate_urls.extend(extract_contract_urls_from_search_text(text))
    except Exception:
        candidate_urls = []
    cleaned_urls = []
    seen = set()
    for url in candidate_urls:
        url = url.replace('&amp;', '&')
        url = url.replace('webapps.dot.illinois.gov', 'webapps1.dot.illinois.gov')
        url = url.strip()
        if 'LbContractDetail/Index' not in url:
            continue
        if url in seen:
            continue
        seen.add(url)
        cleaned_urls.append(url)
        if len(cleaned_urls) >= PUBLIC_SEARCH_MAX_CANDIDATES:
            break
    for url in cleaned_urls:
        try:
            contract_html = get_html(session, url)
            metadata = parse_metadata_from_contract_page(contract_html, url)
            label = metadata.get('item_contract', '')
            if contract_matches(user_job_number, label):
                result = {'label': label, 'url': url, 'letting': metadata.get('letting_date', 'Found by public contract search'), 'letting_url': '', 'page': '', 'source': 'public-contract-search-fast'}
                sqlite_save_contract_result(label, result)
                return result
        except Exception:
            continue
    return None

def find_contract_detail_url(session, job_number):
    """
    Fast lookup path.

    Order:
    1. Direct contract URL, if pasted.
    2. Local SQLite lookup.
    3. Small recent/current IDOT page scan.
    4. Optional tiny public search, off by default.
    5. Optional full archive/index search, off by default.

    The point is to stop the app from hanging during a normal job lookup.
    """
    start_time = time.time()
    original_input = job_number.strip()
    if not original_input:
        raise ValueError('Enter a job number first.')
    if original_input.lower().startswith('http') and 'lbcontractdetail' in original_input.lower():
        html = get_html(session, original_input)
        metadata = parse_metadata_from_contract_page(html, original_input)
        if not metadata.get('item_contract'):
            raise ValueError('The direct URL opened, but the contract number could not be parsed from that page.')
        result = {'label': metadata.get('item_contract', ''), 'url': original_input, 'letting': metadata.get('letting_date', 'Direct URL'), 'letting_url': '', 'page': '', 'source': 'direct-url'}
        sqlite_save_contract_result(result['label'], result)
        return result
    user_job_number = normalize_contract_input(original_input)
    result = sqlite_lookup_contract(user_job_number)
    if result is not None:
        return result
    if sqlite_is_recent_miss(user_job_number):
        raise ValueError(f"Could not find contract '{user_job_number}'. This job number was already searched recently. Try the full item-contract number like 001-62K33, or paste the direct contract detail URL.")
    checked_lettings = 0
    checked_pages = 0
    try:
        letting_links = cached_archive_letting_links_newest_first()
    except Exception:
        letting_links = get_all_archive_letting_links_newest_first(session)
    recent_lettings = letting_links[:RECENT_LETTINGS_FIRST]
    for letting in recent_lettings:
        if search_time_exceeded(start_time):
            break
        checked_lettings += 1
        before = time.time()
        result = letting_page_matches_contract(session=session, letting=letting, user_job_number=user_job_number, max_pages=FAST_PAGES_PER_LETTING, start_time=start_time)
        checked_pages += FAST_PAGES_PER_LETTING
        if result is not None:
            return result
        if time.time() - before > REQUEST_TIMEOUT_SECONDS + 2:
            break
    result = sqlite_lookup_contract(user_job_number)
    if result is not None:
        return result
    if ENABLE_PUBLIC_SEARCH_ON_FAST_LOOKUP and (not search_time_exceeded(start_time)):
        result = cached_public_contract_search(user_job_number)
        if result is not None:
            return result
    if ENABLE_FULL_ARCHIVE_ON_FAST_LOOKUP and (not search_time_exceeded(start_time)):
        for letting in letting_links[RECENT_LETTINGS_FIRST:]:
            if search_time_exceeded(start_time):
                break
            checked_lettings += 1
            result = letting_page_matches_contract(session=session, letting=letting, user_job_number=user_job_number, max_pages=1, start_time=start_time)
            checked_pages += 1
            if result is not None:
                return result
    sqlite_record_miss(user_job_number)
    raise ValueError(f"Could not find contract '{user_job_number}' quickly. Fast lookup checked {checked_lettings} recent/current letting page group(s) and about {checked_pages} page(s). I stopped before running the slow full archive search. Try the full item-contract number like 001-62K33, or paste the direct contract detail URL.")

def unit_regex():
    escaped_units = sorted([re.escape(u) for u in UNIT_WORDS], key=len, reverse=True)
    return '(' + '|'.join(escaped_units) + ')'

def normalize_unit(unit):
    unit = clean_line(unit).upper()
    replacements = {'CUYD': 'CU YD', 'SQYD': 'SQ YD', 'SQFT': 'SQ FT', 'LSUM': 'L SUM'}
    return replacements.get(unit, unit)

def normalize_pay_item_code(value):
    return clean_line(value).upper()

def get_pay_item_report_url(contract_url, html):
    soup = make_soup(html)
    for a in soup.find_all('a', href=True):
        text = clean_line(a.get_text(' ')).lower()
        href = a['href']
        href_lower = href.lower()
        if 'pay item report' in text or 'getpayitemexcelfile' in href_lower:
            return urljoin(contract_url, href)
    return ''

def normalize_pay_item_df(df):
    rename = {}
    for col in df.columns:
        low = str(col).lower().strip()
        if 'pay item' in low and '#' in low or low in ['pay item', 'item #', 'item number', 'item']:
            rename[col] = 'item_code'
        elif 'uom' in low:
            rename[col] = 'unit'
        elif low in ['unit', 'units']:
            rename[col] = 'unit'
        elif 'description' in low:
            rename[col] = 'item_description'
        elif 'quantity' in low or low == 'qty':
            rename[col] = 'quantity'
        elif 'unit price' in low or 'price' in low:
            rename[col] = 'unit_price'
    df = df.rename(columns=rename)
    needed = ['item_code', 'unit', 'item_description', 'quantity', 'unit_price']
    for col in needed:
        if col not in df.columns:
            df[col] = ''
    df = df[needed]
    df['item_code'] = df['item_code'].astype(str).str.strip().str.upper()
    df['unit'] = df['unit'].astype(str).str.strip().str.upper().apply(normalize_unit)
    df['item_description'] = df['item_description'].astype(str).str.strip()
    df['quantity'] = df['quantity'].astype(str).str.strip()
    df['unit_price'] = df['unit_price'].astype(str).str.strip()
    for col in needed:
        df[col] = df[col].replace('nan', '')
    df = df[df['item_code'].str.match('^[A-Z]?\\d{6,8}[A-Z]?$', na=False)]
    df = df.drop_duplicates(subset=['item_code'], keep='first')
    df = df.reset_index(drop=True)
    return df

def parse_pay_items_from_html_tables(text):
    try:
        tables = pd.read_html(StringIO(text))
    except Exception:
        return pd.DataFrame()
    for table in tables:
        table = table.copy()
        if isinstance(table.columns, pd.MultiIndex):
            table.columns = [' '.join([clean_line(x) for x in col if clean_line(x)]) for col in table.columns]
        else:
            table.columns = [clean_line(c) for c in table.columns]
        normalized = normalize_pay_item_df(table)
        if not normalized.empty:
            return normalized
    return pd.DataFrame()

def parse_pay_items_from_row_tokens(text):
    soup = make_soup(text)
    lines = [clean_line(x) for x in soup.get_text('\n').splitlines() if clean_line(x)]
    rows = []
    current = []
    item_start = re.compile(f'^(?P<item_code>[A-Z]?\\d{{6,8}}[A-Z]?)\\s+(?P<unit>{unit_regex()})\\b', re.IGNORECASE)
    for line in lines:
        if line.lower().startswith('no more pay items'):
            break
        if item_start.search(line):
            if current:
                rows.append(' '.join(current))
                current = []
            current.append(line)
        elif current:
            current.append(line)
    if current:
        rows.append(' '.join(current))
    return parse_pay_item_row_strings(rows)

def parse_pay_item_row_strings(row_strings):
    parsed_rows = []
    row_pattern = re.compile(f'^(?P<item_code>[A-Z]?\\d{{6,8}}[A-Z]?)\\s+(?P<unit>{unit_regex()})\\s+(?P<body>.+)$', re.IGNORECASE)
    for row in row_strings:
        row = clean_line(row)
        match = row_pattern.match(row)
        if not match:
            continue
        item_code = normalize_pay_item_code(match.group('item_code'))
        unit = normalize_unit(match.group('unit'))
        body = clean_line(match.group('body'))
        body = re.sub('\\b(Base items|Specialty items|Non-bid items)\\b', ' ', body, flags=re.IGNORECASE)
        body = clean_line(body)
        number_matches = list(re.finditer('\\$?\\d{1,3}(?:,\\d{3})*(?:\\.\\d+)?|\\$?\\d+(?:\\.\\d+)?', body))
        if not number_matches:
            continue
        unit_price = ''
        quantity = ''
        last_number = number_matches[-1].group(0)
        if last_number.startswith('$'):
            unit_price = last_number
            if len(number_matches) < 2:
                continue
            quantity = number_matches[-2].group(0)
            desc_end = number_matches[-2].start()
        else:
            quantity = last_number
            desc_end = number_matches[-1].start()
        description = clean_line(body[:desc_end])
        if not description:
            continue
        parsed_rows.append({'item_code': item_code, 'unit': unit, 'item_description': description, 'quantity': quantity, 'unit_price': unit_price})
    if not parsed_rows:
        return pd.DataFrame()
    return normalize_pay_item_df(pd.DataFrame(parsed_rows))

def parse_pay_items_from_flat_text(text):
    soup = make_soup(text)
    visible_text = soup.get_text(' ')
    visible_text = visible_text.replace('\xa0', ' ')
    visible_text = re.sub('\\s+', ' ', visible_text).strip()
    lower_text = visible_text.lower()
    starts = [lower_text.find('pay item #'), lower_text.find('pay item'), lower_text.find('item #')]
    starts = [x for x in starts if x != -1]
    if starts:
        visible_text = visible_text[min(starts):]
    end = visible_text.lower().find('no more pay items')
    if end != -1:
        visible_text = visible_text[:end]
    item_start_pattern = re.compile(f'\\b(?P<item_code>[A-Z]?\\d{{6,8}}[A-Z]?)\\s+(?P<unit>{unit_regex()})\\b', re.IGNORECASE)
    matches = list(item_start_pattern.finditer(visible_text))
    if not matches:
        return pd.DataFrame()
    row_strings = []
    for index, match in enumerate(matches):
        start = match.start()
        if index + 1 < len(matches):
            end = matches[index + 1].start()
        else:
            end = len(visible_text)
        row_strings.append(clean_line(visible_text[start:end]))
    return parse_pay_item_row_strings(row_strings)

def parse_pay_items_from_tab_or_csv_text(text):
    for sep in ['\t', ',', ';', '|']:
        try:
            df = pd.read_csv(StringIO(text), sep=sep, engine='python')
        except Exception:
            continue
        normalized = normalize_pay_item_df(df)
        if not normalized.empty:
            return normalized
    return pd.DataFrame()

def find_table_inside_raw_excel(raw_df):
    for row_index in range(len(raw_df)):
        row_values = list(raw_df.iloc[row_index].values)
        joined = ' '.join([clean_line(x).lower() for x in row_values])
        has_item = 'pay item' in joined or 'item #' in joined or 'item number' in joined
        has_uom = 'uom' in joined or 'unit' in joined
        has_desc = 'description' in joined
        has_qty = 'quantity' in joined or 'qty' in joined
        if has_item and has_uom and has_desc and has_qty:
            headers = [clean_line(x) for x in row_values]
            data = raw_df.iloc[row_index + 1:].copy()
            data.columns = headers
            return data
    return pd.DataFrame()

def parse_pay_items_from_excel_bytes(content):
    """
    Parse the Pay Item Excel file with the least retry work possible.

    The old version tried multiple engines and repeatedly reread the same file,
    which made the app feel like search was still running even after the
    contract page had already been found.
    """
    engines = ['openpyxl', None]
    if content[:8].startswith(b'\xd0\xcf\x11\xe0'):
        engines.append('xlrd')
    for engine in engines:
        try:
            file_data = io.BytesIO(content)
            if engine is None:
                sheets = pd.read_excel(file_data, sheet_name=None, header=None)
            else:
                sheets = pd.read_excel(file_data, sheet_name=None, header=None, engine=engine)
            for sheet_name, raw_df in sheets.items():
                df = find_table_inside_raw_excel(raw_df)
                normalized = normalize_pay_item_df(df)
                if not normalized.empty:
                    return normalized
        except Exception:
            continue
    return pd.DataFrame()

def parse_pay_items_from_any_text(text):
    parsers = [parse_pay_items_from_html_tables, parse_pay_items_from_row_tokens, parse_pay_items_from_flat_text, parse_pay_items_from_tab_or_csv_text]
    for parser in parsers:
        try:
            pay_items = parser(text)
            if not pay_items.empty:
                return pay_items
        except Exception:
            continue
    return pd.DataFrame()

def looks_like_excel_response(response, content):
    content_type = clean_line(response.headers.get('Content-Type', '')).lower()
    content_disposition = clean_line(response.headers.get('Content-Disposition', '')).lower()
    if 'excel' in content_type or 'spreadsheet' in content_type:
        return True
    if '.xlsx' in content_disposition or '.xls' in content_disposition:
        return True
    if content.startswith(b'PK') or content[:8].startswith(b'\xd0\xcf\x11\xe0'):
        return True
    return False

def parse_pay_items_from_pay_item_report(session, contract_url, html):
    pay_item_url = get_pay_item_report_url(contract_url, html)
    if not pay_item_url:
        return pd.DataFrame()
    response = get_response(session, pay_item_url)
    content = response.content
    if not content:
        return pd.DataFrame()
    if looks_like_excel_response(response, content):
        pay_items = parse_pay_items_from_excel_bytes(content)
        if not pay_items.empty:
            return pay_items
    text = decode_content(content)
    pay_items = parse_pay_items_from_any_text(text)
    if not pay_items.empty:
        return pay_items
    return parse_pay_items_from_excel_bytes(content)

def parse_pay_items_from_contract_page_text(html):
    return parse_pay_items_from_any_text(html)

@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def fetch_idot_job(job_number):
    session = make_session()
    match = find_contract_detail_url(session, job_number)
    html = get_html(session, match['url'])
    metadata = parse_metadata_from_contract_page(html, match['url'])
    if not metadata.get('item_contract'):
        metadata['item_contract'] = match.get('label', '')
    pay_items = parse_pay_items_from_pay_item_report(session, match['url'], html)
    if pay_items.empty:
        pay_items = parse_pay_items_from_contract_page_text(html)
    if pay_items.empty:
        raise ValueError('Found the contract page, but could not extract pay items from the Pay Item Report or page text.')
    return (metadata, pay_items, match)
import shutil
import subprocess
import tempfile
from datetime import date as DateClass
PDF_ROW_COUNT = 6
WEATHER_OPTIONS = ['Sunny', 'Cloudy', 'Light Rain', 'Normal Rain', 'Heavy Rain', 'Snow']

def get_today_default():
    return DateClass.today()

def format_report_date(value):
    if hasattr(value, 'strftime'):
        return value.strftime('%-m/%-d/%Y') if hasattr(value, 'strftime') else str(value)
    return clean_line(value)

def format_pdf_filename(contract_number):
    contract_number = clean_line(contract_number) or 'IDOT'
    safe = re.sub('[^A-Za-z0-9_-]+', '_', contract_number).strip('_')
    return f'{safe}_IDR.pdf'

def format_xlsx_filename(contract_number):
    contract_number = clean_line(contract_number) or 'IDOT'
    safe = re.sub('[^A-Za-z0-9_-]+', '_', contract_number).strip('_')
    return f'{safe}_IDR_filled.xlsx'

def dataframe_records_by_code(pay_items):
    records = {}
    if pay_items is None or pay_items.empty:
        return records
    for _, row in pay_items.iterrows():
        code = normalize_pay_item_code(row.get('item_code', ''))
        if code and code not in records:
            records[code] = {'item_code': code, 'item_description': clean_line(row.get('item_description', '')), 'unit': normalize_unit(row.get('unit', '')), 'plan_quantity': clean_line(row.get('quantity', '')), 'unit_price': clean_line(row.get('unit_price', '')), 'is_custom': False}
    return records

def dataframe_records_by_description(pay_items):
    records = {}
    if pay_items is None or pay_items.empty:
        return records
    for _, row in pay_items.iterrows():
        desc = clean_line(row.get('item_description', ''))
        if desc and desc not in records:
            code = normalize_pay_item_code(row.get('item_code', ''))
            records[desc] = {'item_code': code, 'item_description': desc, 'unit': normalize_unit(row.get('unit', '')), 'plan_quantity': clean_line(row.get('quantity', '')), 'unit_price': clean_line(row.get('unit_price', '')), 'is_custom': False}
    return records

def parse_number(value):
    value = clean_line(value)
    if not value:
        return None
    value = value.replace(',', '').replace('$', '')
    value = re.sub('[^0-9.\\-]', '', value)
    if not value or value in ['.', '-', '-.']:
        return None
    try:
        return float(value)
    except Exception:
        return None

def get_quantity_status(quantity, plan_quantity):
    entered = parse_number(quantity)
    permitted = parse_number(plan_quantity)
    if entered is None or permitted is None or permitted <= 0:
        return {'status': '', 'ratio': None, 'color': ''}
    ratio = entered / permitted
    if entered > permitted:
        return {'status': f'OVER ({ratio:.0%})', 'ratio': ratio, 'color': '#ffc7ce'}
    if ratio >= 0.75:
        return {'status': f'Close ({ratio:.0%})', 'ratio': ratio, 'color': '#ffeb9c'}
    return {'status': f'OK ({ratio:.0%})', 'ratio': ratio, 'color': '#c6efce'}

def get_pay_item_options(pay_items):
    if pay_items is None or pay_items.empty:
        return (['', 'Custom / Manual'], ['', 'Custom / Manual'])
    codes = ['']
    descriptions = ['']
    for _, row in pay_items.iterrows():
        code = normalize_pay_item_code(row.get('item_code', ''))
        desc = clean_line(row.get('item_description', ''))
        if code and code not in codes:
            codes.append(code)
        if desc and desc not in descriptions:
            descriptions.append(desc)
    codes.append('Custom / Manual')
    descriptions.append('Custom / Manual')
    return (codes, descriptions)

def get_pay_item_by_code(pay_items, code):
    return dataframe_records_by_code(pay_items).get(normalize_pay_item_code(code))

def get_pay_item_by_description(pay_items, description):
    return dataframe_records_by_description(pay_items).get(clean_line(description))

def row_key(row_index, field):
    return f'idr_row_{row_index}_{field}'

def header_key(field):
    return f'idr_header_{field}'

def clear_idr_row_state():
    for row_index in range(PDF_ROW_COUNT):
        for field in ['item_code', 'item_description', 'custom_code', 'custom_description', 'location', 'quantity', 'unit', 'custom_unit', 'plan_quantity', 'unit_price', 'evidence', 'is_custom']:
            st.session_state.pop(row_key(row_index, field), None)

def ensure_row_defaults(row_index):
    defaults = {'item_code': '', 'item_description': '', 'custom_code': '', 'custom_description': '', 'location': '', 'quantity': '', 'unit': '', 'custom_unit': '', 'plan_quantity': '', 'unit_price': '', 'evidence': '', 'is_custom': False}
    for field, value in defaults.items():
        st.session_state.setdefault(row_key(row_index, field), value)

def set_row_from_official_item(row_index, item):
    st.session_state[row_key(row_index, 'item_code')] = clean_line(item.get('item_code', ''))
    st.session_state[row_key(row_index, 'item_description')] = clean_line(item.get('item_description', ''))
    st.session_state[row_key(row_index, 'unit')] = normalize_unit(item.get('unit', ''))
    st.session_state[row_key(row_index, 'plan_quantity')] = clean_line(item.get('plan_quantity', ''))
    st.session_state[row_key(row_index, 'unit_price')] = clean_line(item.get('unit_price', ''))
    st.session_state[row_key(row_index, 'evidence')] = get_evidence_of_material_inspection(item.get('item_code', ''), item.get('item_description', ''))
    st.session_state[row_key(row_index, 'is_custom')] = False

def on_item_code_change(row_index, pay_items):
    code = st.session_state.get(row_key(row_index, 'item_code'), '')
    if code == 'Custom / Manual':
        st.session_state[row_key(row_index, 'item_description')] = 'Custom / Manual'
        st.session_state[row_key(row_index, 'is_custom')] = True
        st.session_state[row_key(row_index, 'plan_quantity')] = ''
        st.session_state[row_key(row_index, 'unit_price')] = ''
        st.session_state[row_key(row_index, 'evidence')] = ''
        return
    if not code:
        st.session_state[row_key(row_index, 'item_description')] = ''
        st.session_state[row_key(row_index, 'unit')] = ''
        st.session_state[row_key(row_index, 'plan_quantity')] = ''
        st.session_state[row_key(row_index, 'unit_price')] = ''
        st.session_state[row_key(row_index, 'is_custom')] = False
        return
    item = get_pay_item_by_code(pay_items, code)
    if item:
        set_row_from_official_item(row_index, item)

def on_item_description_change(row_index, pay_items):
    description = st.session_state.get(row_key(row_index, 'item_description'), '')
    if description == 'Custom / Manual':
        st.session_state[row_key(row_index, 'item_code')] = 'Custom / Manual'
        st.session_state[row_key(row_index, 'is_custom')] = True
        st.session_state[row_key(row_index, 'plan_quantity')] = ''
        st.session_state[row_key(row_index, 'unit_price')] = ''
        st.session_state[row_key(row_index, 'evidence')] = ''
        return
    if not description:
        st.session_state[row_key(row_index, 'item_code')] = ''
        st.session_state[row_key(row_index, 'unit')] = ''
        st.session_state[row_key(row_index, 'plan_quantity')] = ''
        st.session_state[row_key(row_index, 'unit_price')] = ''
        st.session_state[row_key(row_index, 'is_custom')] = False
        return
    item = get_pay_item_by_description(pay_items, description)
    if item:
        set_row_from_official_item(row_index, item)

def on_custom_row_change(row_index):
    if st.session_state.get(row_key(row_index, 'is_custom'), False):
        st.session_state[row_key(row_index, 'unit')] = normalize_unit(st.session_state.get(row_key(row_index, 'custom_unit'), ''))

def get_row_for_output(row_index):
    is_custom = bool(st.session_state.get(row_key(row_index, 'is_custom'), False))
    if is_custom:
        return {'item_code': normalize_pay_item_code(st.session_state.get(row_key(row_index, 'custom_code'), '')), 'item_description': clean_line(st.session_state.get(row_key(row_index, 'custom_description'), '')), 'location': clean_line(st.session_state.get(row_key(row_index, 'location'), '')), 'quantity': clean_line(st.session_state.get(row_key(row_index, 'quantity'), '')), 'unit': normalize_unit(st.session_state.get(row_key(row_index, 'custom_unit'), '')), 'plan_quantity': '', 'unit_price': '', 'evidence': '', 'is_custom': True}
    return {'item_code': clean_line(st.session_state.get(row_key(row_index, 'item_code'), '')), 'item_description': clean_line(st.session_state.get(row_key(row_index, 'item_description'), '')), 'location': clean_line(st.session_state.get(row_key(row_index, 'location'), '')), 'quantity': clean_line(st.session_state.get(row_key(row_index, 'quantity'), '')), 'unit': normalize_unit(st.session_state.get(row_key(row_index, 'unit'), '')), 'plan_quantity': clean_line(st.session_state.get(row_key(row_index, 'plan_quantity'), '')), 'unit_price': clean_line(st.session_state.get(row_key(row_index, 'unit_price'), '')), 'evidence': get_evidence_of_material_inspection(st.session_state.get(row_key(row_index, 'item_code'), ''), st.session_state.get(row_key(row_index, 'item_description'), '')), 'is_custom': False}
STANDARD_REMARKS_INSTRUCTION = '(e.g. instruction to Contractor, special problems, sketches with dimensions for final measurements, computations, number of persons working, hours worked) Use reverse side, if needed.'

def selected_item_codes(rows):
    """Return every nonblank item code shown on the IDR, including repeats."""
    codes = []
    for row in rows or []:
        code = normalize_pay_item_code(row.get('item_code', ''))
        if not code or code == 'CUSTOM / MANUAL':
            continue
        codes.append(code)
    return codes

def quantity_status_badge_html(quantity, plan_quantity):
    status = get_quantity_status(quantity, plan_quantity)
    if not status['status']:
        return "<div class='qty-badge qty-empty'>-</div>"
    return f"<div class='qty-badge' style='background:{status['color']};'>{status['status']}</div>"

def build_idr_header_form():
    st.subheader('IDR Header / Top of Form')
    st.caption('Fill these boxes exactly like the top and signature sections of the Excel IDR. The labels shown here explain where each value prints on the final PDF.')
    with st.expander('Show field guide', expanded=False):
        st.markdown('\n            - **Date** → prints in the top-left date box. The dates beside Inspected/Measured/Calculated only print when initials are entered.\n            - **Contractor or Sub.** → prints on the Contractor/Subcontractor line.\n            - **Weather** → select a standard option or choose Custom / Manual and type your own weather.\n            - **Inspected by / Measured by / Calculated by** → prints in the signature/initial boxes on the right side of the form.\n            - **This is** → checks either Estimated Progress Measurement or Final Field Measurement.\n            - **Item no.** → all selected pay-item codes print automatically beside the selected measurement checkbox.\n            - **Remarks** → prints in the expanded Remarks box under the measurement section.\n            ')
    st.markdown('**Top form fields**')
    row1 = st.columns([1.0, 2.0, 1.4, 1.1, 1.1, 1.1])
    with row1[0]:
        idr_date = st.date_input('Date', value=get_today_default(), key=header_key('date'))
    with row1[1]:
        contractor = st.text_input('Contractor or Sub.', key=header_key('contractor'))
    with row1[2]:
        weather_choice = st.selectbox('Weather', [''] + WEATHER_OPTIONS + ['Custom / Manual'], key=header_key('weather'))
        if weather_choice == 'Custom / Manual':
            weather = st.text_input('Custom weather', key=header_key('custom_weather'), placeholder='Enter weather conditions')
        else:
            weather = weather_choice
    with row1[3]:
        inspected_by = st.text_input('Inspected by', key=header_key('inspected_by'))
    with row1[4]:
        measured_by = st.text_input('Measured by', key=header_key('measured_by'))
    with row1[5]:
        calculated_by = st.text_input('Calculated by', key=header_key('calculated_by'))
    st.markdown('**Measurement and remarks fields**')
    row2 = st.columns([1.4, 4.6])
    with row2[0]:
        measurement_type = st.selectbox('This is', ['', 'Estimated progress measurement', 'Final field measurement'], key=header_key('measurement_type'))
        st.caption('The selected pay-item codes are inserted automatically in the item-no. line.')
    with row2[1]:
        remarks = st.text_area('Remarks', height=70, key=header_key('remarks'))
    st.markdown('**COGO area calculation statement**')
    cogo_cols = st.columns([1.5, 1.0, 3.5])
    with cogo_cols[0]:
        cogo_statement_option = st.selectbox('Include COGO statement', ['No', 'Yes'], key=header_key('cogo_statement_option'))
    with cogo_cols[1]:
        cogo_version_year = st.text_input('Trimble Access version year', key=header_key('cogo_version_year'), disabled=cogo_statement_option != 'Yes', placeholder='e.g. 2026')
    with cogo_cols[2]:
        st.caption('When selected, the statement prints above the typed remarks on the IDR.')
    return {'date': idr_date, 'contractor': contractor, 'weather': weather, 'inspected_by': inspected_by, 'measured_by': measured_by, 'calculated_by': calculated_by, 'measurement_type': measurement_type, 'remarks': remarks, 'cogo_statement_option': cogo_statement_option, 'cogo_version_year': cogo_version_year}

def build_idr_rows_form(pay_items):
    st.subheader('IDR Pay Item Table')
    st.caption('Select an official item for automatic item data, or choose Custom / Manual in either the Item Code or Item field. A custom code and custom item box will then appear. Manual rows do not auto-fill evidence or plan quantity.')
    code_options, description_options = get_pay_item_options(pay_items)
    st.markdown('\n        <style>\n        .idr-table-header {font-weight: 700; font-size: 0.82rem; padding: 0.25rem 0; border-bottom: 1px solid #d0d0d0;}\n        .qty-badge {min-height: 36px; border: 1px solid #999; border-radius: 4px; display: flex; align-items: center; justify-content: center; font-weight: 700; font-size: 0.78rem; color: #111; margin-top: 1.70rem;}\n        .qty-empty {background: #f4f4f4; color: #777; font-weight: 400;}\n        </style>\n        ', unsafe_allow_html=True)
    header_cols = st.columns([1.0, 0.55, 2.45, 1.55, 0.75, 0.7, 1.25])
    headers = ['Item Code #', 'Fund', 'Item', 'Location', 'Quantity', 'Unit', 'Plan Quantity Check']
    for col, header in zip(header_cols, headers):
        col.markdown(f"<div class='idr-table-header'>{header}</div>", unsafe_allow_html=True)
    rows = []
    for row_index in range(PDF_ROW_COUNT):
        ensure_row_defaults(row_index)
        row_cols = st.columns([1.0, 0.55, 2.45, 1.55, 0.75, 0.7, 1.25])
        current_code = st.session_state.get(row_key(row_index, 'item_code'), '')
        current_desc = st.session_state.get(row_key(row_index, 'item_description'), '')
        if current_code not in code_options:
            st.session_state[row_key(row_index, 'item_code')] = ''
        if current_desc not in description_options:
            st.session_state[row_key(row_index, 'item_description')] = ''
        with row_cols[0]:
            st.selectbox(f'Row {row_index + 1} Item Code', code_options, key=row_key(row_index, 'item_code'), on_change=on_item_code_change, args=(row_index, pay_items), label_visibility='collapsed')
        with row_cols[1]:
            st.text_input(f'Row {row_index + 1} Fund', key=row_key(row_index, 'fund_code'), label_visibility='collapsed')
        with row_cols[2]:
            st.selectbox(f'Row {row_index + 1} Item Description', description_options, key=row_key(row_index, 'item_description'), on_change=on_item_description_change, args=(row_index, pay_items), label_visibility='collapsed')
        is_custom = bool(st.session_state.get(row_key(row_index, 'is_custom'), False))
        if is_custom:
            with row_cols[0]:
                st.text_input(f'Row {row_index + 1} Custom Code', key=row_key(row_index, 'custom_code'), placeholder='Custom ID', label_visibility='collapsed')
            with row_cols[2]:
                st.text_input(f'Row {row_index + 1} Custom Description', key=row_key(row_index, 'custom_description'), placeholder='Custom item description', label_visibility='collapsed')
        with row_cols[3]:
            st.text_input(f'Row {row_index + 1} Location', key=row_key(row_index, 'location'), label_visibility='collapsed')
        with row_cols[4]:
            st.text_input(f'Row {row_index + 1} Quantity', key=row_key(row_index, 'quantity'), label_visibility='collapsed')
        with row_cols[5]:
            if is_custom:
                custom_code_value = clean_line(st.session_state.get(row_key(row_index, 'custom_code'), ''))
                custom_description_value = clean_line(st.session_state.get(row_key(row_index, 'custom_description'), ''))
                custom_identity_entered = bool(custom_code_value or custom_description_value)
                if not custom_identity_entered:
                    st.session_state[row_key(row_index, 'custom_unit')] = ''
                st.text_input(f'Row {row_index + 1} Unit', key=row_key(row_index, 'custom_unit'), placeholder='Unit' if custom_identity_entered else 'Enter custom item/code first', label_visibility='collapsed', disabled=not custom_identity_entered, on_change=on_custom_row_change, args=(row_index,))
            else:
                st.text_input(f'Row {row_index + 1} Unit', key=row_key(row_index, 'unit'), label_visibility='collapsed', disabled=True)
        row = get_row_for_output(row_index)
        row['fund_code'] = clean_line(st.session_state.get(row_key(row_index, 'fund_code'), ''))
        rows.append(row)
        with row_cols[6]:
            st.markdown(quantity_status_badge_html(row.get('quantity', ''), row.get('plan_quantity', '')), unsafe_allow_html=True)
    return rows

def safe_set(ws, cell, value):
    """Write to a cell safely, including merged cells."""
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            top_left_cell = merged_range.coord.split(':')[0]
            ws[top_left_cell] = value
            return
    ws[cell] = value

def get_merged_anchor_cell(ws, cell_address):
    for merged_range in ws.merged_cells.ranges:
        if cell_address in merged_range:
            return merged_range.coord.split(':')[0]
    return cell_address

def get_description_font_size(description):
    description_length = len(clean_line(description))
    for max_length, font_size in DESCRIPTION_TEXT_FIT_RULES:
        if description_length <= max_length:
            return font_size
    return 6

def make_font_with_size(original_font, size):
    return Font(name=original_font.name, sz=size, b=original_font.b, i=original_font.i, vertAlign=original_font.vertAlign, underline=original_font.underline, strike=original_font.strike, color=copy(original_font.color), scheme=original_font.scheme, family=original_font.family, charset=original_font.charset, outline=original_font.outline, shadow=original_font.shadow, condense=original_font.condense, extend=original_font.extend)

def get_text_for_cell(ws, cell_address):
    anchor_address = get_merged_anchor_cell(ws, cell_address)
    value = ws[anchor_address].value
    return '' if value is None else str(value)

def format_item_description_cells(ws):
    for row in range(13, 19):
        visible_cell_address = f'D{row}'
        anchor_cell_address = get_merged_anchor_cell(ws, visible_cell_address)
        anchor_cell = ws[anchor_cell_address]
        description_text = get_text_for_cell(ws, visible_cell_address)
        font_size = get_description_font_size(description_text)
        current_alignment = copy(anchor_cell.alignment)
        anchor_cell.alignment = Alignment(horizontal=current_alignment.horizontal or 'left', vertical='top', text_rotation=current_alignment.text_rotation, wrap_text=True, shrink_to_fit=False, indent=current_alignment.indent, relativeIndent=current_alignment.relativeIndent, justifyLastLine=current_alignment.justifyLastLine, readingOrder=current_alignment.readingOrder)
        anchor_cell.font = make_font_with_size(anchor_cell.font, font_size)

def format_quantity_cells(ws):
    """
    Give the quantity/unit cells a little breathing room so the left edge
    of the number does not get clipped in the exported PDF.

    The template stores the visible quantity box around H13:H18. LibreOffice
    PDF export can render left-aligned text too tight against the border,
    so we keep the same Excel layout but center the value vertically and
    add a small indent.
    """
    for row in range(13, 19):
        visible_cell_address = f'H{row}'
        anchor_cell_address = get_merged_anchor_cell(ws, visible_cell_address)
        anchor_cell = ws[anchor_cell_address]
        current_alignment = copy(anchor_cell.alignment)
        anchor_cell.alignment = Alignment(horizontal='left', vertical=current_alignment.vertical or 'center', text_rotation=current_alignment.text_rotation, wrap_text=current_alignment.wrap_text, shrink_to_fit=False, indent=1, relativeIndent=current_alignment.relativeIndent, justifyLastLine=current_alignment.justifyLastLine, readingOrder=current_alignment.readingOrder)
        for merged_range in ws.merged_cells.ranges:
            if visible_cell_address in merged_range:
                for cell_row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row, min_col=merged_range.min_col, max_col=merged_range.max_col):
                    for range_cell in cell_row:
                        try:
                            range_cell.alignment = copy(anchor_cell.alignment)
                        except Exception:
                            pass
                break

def unmerge_range_keep_style(ws, range_coord):
    target = None
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == range_coord:
            target = merged_range
            break
    if target is None:
        return
    anchor = ws.cell(target.min_row, target.min_col)
    saved = {'font': copy(anchor.font), 'fill': copy(anchor.fill), 'border': copy(anchor.border), 'alignment': copy(anchor.alignment), 'number_format': anchor.number_format, 'protection': copy(anchor.protection)}
    ws.unmerge_cells(range_coord)
    for row in range(target.min_row, target.max_row + 1):
        for col in range(target.min_col, target.max_col + 1):
            cell = ws.cell(row, col)
            cell.font = copy(saved['font'])
            cell.fill = copy(saved['fill'])
            cell.border = copy(saved['border'])
            cell.alignment = copy(saved['alignment'])
            cell.number_format = saved['number_format']
            cell.protection = copy(saved['protection'])

def increase_entire_form_font_size(ws, points=1.5):
    """Increase every visible font in the printable IDR area by the requested points."""
    for row in ws.iter_rows(min_row=2, max_row=34, min_col=1, max_col=14):
        for cell in row:
            try:
                current_size = cell.font.sz
                if current_size is not None:
                    cell.font = make_font_with_size(cell.font, float(current_size) + points)
            except Exception:
                pass

def format_evidence_cells(ws):
    """Wrap and fit automatically populated evidence text in J13:L18."""
    for row in range(13, 19):
        visible_cell_address = f'J{row}'
        anchor_cell_address = get_merged_anchor_cell(ws, visible_cell_address)
        anchor_cell = ws[anchor_cell_address]
        evidence_text = get_text_for_cell(ws, visible_cell_address)
        current_alignment = copy(anchor_cell.alignment)
        anchor_cell.alignment = Alignment(horizontal='left', vertical='top', text_rotation=current_alignment.text_rotation, wrap_text=True, shrink_to_fit=False, indent=0, relativeIndent=current_alignment.relativeIndent, justifyLastLine=current_alignment.justifyLastLine, readingOrder=current_alignment.readingOrder)
        length = len(clean_line(evidence_text))
        if length <= 28:
            size = 9
        elif length <= 60:
            size = 8
        elif length <= 100:
            size = 7
        else:
            size = 6
        anchor_cell.font = make_font_with_size(anchor_cell.font, size)

def prepare_exact_print_layout(wb, ws):
    wb.active = wb.sheetnames.index(ws.title)
    for sheet in wb.worksheets:
        if sheet.title != ws.title:
            sheet.sheet_state = 'hidden'
    ws.print_area = 'A2:N34'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.18
    ws.page_margins.bottom = 0.18
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

def copy_cell_style(source_cell, target_cell):
    """Copy the visible Excel style from one cell to another."""
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)

def rebuild_bottom_section_layout(ws):
    """
    Rebuild the lower part of the IDR without separator lines:
    - rows 19-20: measurement checkboxes and item numbers
    - rows 21-22: permanent Remarks instruction
    - rows 23-24: optional COGO statement, left-aligned with remarks
    - rows 25-33: user's typed remarks
    - row 34: printed date and revision footer
    """
    label_style_source = ws['B21']
    checkbox_style_source = ws['C21']
    measurement_text_style_source = ws['D21']
    remarks_label_style_source = ws['B25']
    remarks_box_style_source = ws['C25']
    footer_left_style_source = ws['B30']
    footer_right_style_source = ws['N30']
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= 34 and merged_range.max_row >= 19:
            try:
                unmerge_range_keep_style(ws, str(merged_range))
            except Exception:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception:
                    pass
    for row in range(19, 35):
        for col in range(1, 15):
            ws.cell(row=row, column=col).value = None
    safe_set(ws, 'B19', 'This is:')
    safe_set(ws, 'C19', '☐')
    safe_set(ws, 'D19', 'an estimated progress measurement')
    safe_set(ws, 'F19', '(item no.:')
    safe_set(ws, 'G19', '')
    safe_set(ws, 'L19', ')')
    safe_set(ws, 'C20', '☐')
    safe_set(ws, 'D20', 'a final field measurement')
    safe_set(ws, 'F20', '(item no.:')
    safe_set(ws, 'G20', '')
    safe_set(ws, 'L20', ')')
    safe_set(ws, 'B21', 'Remarks:')
    safe_set(ws, 'C21', STANDARD_REMARKS_INSTRUCTION)
    safe_set(ws, 'C23', '')
    safe_set(ws, 'C25', '')
    safe_set(ws, 'A34', '')
    safe_set(ws, 'M34', 'BC 628 (Rev. 8/04)')
    for merge in ['D19:E19', 'D20:E20', 'G19:K19', 'G20:K20', 'C21:N22', 'C23:N24', 'C25:N33', 'A34:D34', 'M34:N34']:
        try:
            ws.merge_cells(merge)
        except Exception:
            pass
    copy_cell_style(label_style_source, ws['B19'])
    for addr in ['C19', 'C20']:
        copy_cell_style(checkbox_style_source, ws[addr])
    for addr in ['D19', 'F19', 'G19', 'D20', 'F20', 'G20']:
        copy_cell_style(measurement_text_style_source, ws[addr])
        ws[addr].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=True)

    # Keep both measurement descriptions at exactly the same font size.
    # Without this, Excel/LibreOffice can shrink the longer estimated-progress
    # label more than the final-measurement label.
    shared_measurement_font = make_font_with_size(
        copy(measurement_text_style_source.font), 10.5
    )
    for addr in ['D19', 'D20', 'F19', 'F20', 'G19', 'G20']:
        ws[addr].font = copy(shared_measurement_font)
        ws[addr].alignment = Alignment(
            horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=False
        )
    for addr in ['L19', 'L20']:
        copy_cell_style(measurement_text_style_source, ws[addr])
        ws[addr].alignment = Alignment(horizontal='right', vertical='center', wrap_text=False, shrink_to_fit=False)
    copy_cell_style(remarks_label_style_source, ws['B21'])
    for anchor_addr in ['C21', 'C23', 'C25']:
        anchor = ws[anchor_addr]
        copy_cell_style(remarks_box_style_source, anchor)
        anchor.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, shrink_to_fit=False)
        anchor.border = Border()
    ws['C21'].font = make_font_with_size(ws['C21'].font, 10)
    ws['C23'].font = make_font_with_size(ws['C23'].font, 10)
    for row in range(19, 34):
        for col in range(2, 15):
            ws.cell(row=row, column=col).border = Border()
    writing_line = Border(bottom=Side(style='thin', color='000000'))
    for row in (19, 20):
        for col in range(7, 13):
            ws.cell(row=row, column=col).border = writing_line
    for row in range(21, 23):
        ws.row_dimensions[row].height = 15
    for row in range(23, 25):
        ws.row_dimensions[row].height = 15
    for row in range(25, 34):
        ws.row_dimensions[row].height = 15
    ws.row_dimensions[34].height = 16
    copy_cell_style(footer_left_style_source, ws['A34'])
    ws['A34'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    copy_cell_style(footer_right_style_source, ws['M34'])
    ws['M34'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

def ensure_contractor_value_area(ws):
    """Keep the Contractor/Sub label intact and provide a reliable value area at D8:F8."""
    containing = None
    for merged_range in list(ws.merged_cells.ranges):
        if 'D8' in merged_range:
            containing = str(merged_range)
            break
    if containing and (not containing.startswith('D8:')):
        try:
            unmerge_range_keep_style(ws, containing)
        except Exception:
            try:
                ws.unmerge_cells(containing)
            except Exception:
                pass
    try:
        already = any(('D8' in r and str(r).startswith('D8:') for r in ws.merged_cells.ranges))
        if not already:
            ws.merge_cells('D8:F8')
    except Exception:
        pass

def clear_exact_idr_values(ws):
    for cell in ['C6', 'D8', 'C10', 'G6', 'H6', 'G7', 'H7', 'G8', 'H8', 'G9', 'H9', 'L2', 'L3', 'L4', 'L5', 'L6', 'L7', 'L8', 'C21', 'C23', 'C25', 'D19', 'F19', 'G19', 'L19', 'D20', 'F20', 'G20', 'L20', 'A34', 'M34']:
        safe_set(ws, cell, '')
    safe_set(ws, 'C19', '☐')
    safe_set(ws, 'C20', '☐')
    safe_set(ws, 'D19', 'an estimated progress measurement')
    safe_set(ws, 'F19', '(item no.:')
    safe_set(ws, 'G19', '')
    safe_set(ws, 'L19', ')')
    safe_set(ws, 'D20', 'a final field measurement')
    safe_set(ws, 'F20', '(item no.:')
    safe_set(ws, 'G20', '')
    safe_set(ws, 'L20', ')')
    safe_set(ws, 'M34', 'BC 628 (Rev. 8/04)')
    for row in range(13, 19):
        for col in ['B', 'C', 'D', 'F', 'H', 'I', 'J', 'M']:
            safe_set(ws, f'{col}{row}', '')

def format_short_contract_number(value):
    """Return only the five-character IDOT contract suffix, e.g. 62V36."""
    value = normalize_contract_input(value)
    if not value:
        return ''
    if '-' in value:
        value = value.split('-')[-1]
    return value[-5:]

def fill_exact_idr_workbook(metadata, idr_info, rows):
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f'Template file not found. Put IDR_Template.xlsx in the same folder as this app file. Expected: {TEMPLATE_PATH}')
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['IDR_Form'] if 'IDR_Form' in wb.sheetnames else wb.active
    for merged in ['G6:H6', 'G7:H7', 'G8:H8', 'G9:H9']:
        unmerge_range_keep_style(ws, merged)
    rebuild_bottom_section_layout(ws)
    ensure_contractor_value_area(ws)
    clear_exact_idr_values(ws)
    report_date = format_report_date(idr_info.get('date'))
    safe_set(ws, 'C6', report_date)
    safe_set(ws, 'A34', f'Printed {report_date}')
    safe_set(ws, 'M34', 'BC 628 (Rev. 8/04)')
    safe_set(ws, 'D8', clean_line(idr_info.get('contractor', '')))
    safe_set(ws, 'C10', idr_info.get('weather', ''))
    inspected_by = clean_line(idr_info.get('inspected_by', ''))
    measured_by = clean_line(idr_info.get('measured_by', ''))
    calculated_by = clean_line(idr_info.get('calculated_by', ''))
    safe_set(ws, 'G6', inspected_by)
    safe_set(ws, 'H6', report_date if inspected_by else '')
    safe_set(ws, 'G7', measured_by)
    safe_set(ws, 'H7', report_date if measured_by else '')
    safe_set(ws, 'G8', calculated_by)
    safe_set(ws, 'H8', report_date if calculated_by else '')
    safe_set(ws, 'G9', '')
    safe_set(ws, 'H9', '')
    safe_set(ws, 'L2', metadata.get('county', ''))
    safe_set(ws, 'L3', metadata.get('key_route', ''))
    safe_set(ws, 'L4', metadata.get('marked_route', ''))
    safe_set(ws, 'L5', metadata.get('district', ''))
    safe_set(ws, 'L6', format_short_contract_number(metadata.get('item_contract', '')))
    safe_set(ws, 'L7', metadata.get('state_job', ''))
    safe_set(ws, 'L8', metadata.get('federal_project', ''))
    measurement_type = clean_line(idr_info.get('measurement_type', ''))
    automatic_item_numbers = ', '.join(selected_item_codes(rows))
    estimated_numbers = automatic_item_numbers if measurement_type == 'Estimated progress measurement' else ''
    final_numbers = automatic_item_numbers if measurement_type == 'Final field measurement' else ''
    safe_set(ws, 'D19', 'an estimated progress measurement')
    safe_set(ws, 'F19', '(item no.:')
    safe_set(ws, 'G19', estimated_numbers)
    safe_set(ws, 'L19', ')')
    safe_set(ws, 'D20', 'a final field measurement')
    safe_set(ws, 'F20', '(item no.:')
    safe_set(ws, 'G20', final_numbers)
    safe_set(ws, 'L20', ')')
    for addr in ['G19', 'G20']:
        ws[addr].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=True)
    if measurement_type == 'Estimated progress measurement':
        safe_set(ws, 'C19', '☒')
    elif measurement_type == 'Final field measurement':
        safe_set(ws, 'C20', '☒')
    safe_set(ws, 'C21', STANDARD_REMARKS_INSTRUCTION)
    cogo_statement = ''
    if clean_line(idr_info.get('cogo_statement_option', '')) == 'Yes':
        cogo_year = clean_line(idr_info.get('cogo_version_year', ''))
        version_text = f' {cogo_year}' if cogo_year else ''
        cogo_statement = f'Used Cogo Area Calc Tool Trimble Access Version{version_text} Area Calculated from a list of points shot around the perimeter (attached Area Calculation, pointlist, and coordinates measured quantity compares to plan quantity.)'
    safe_set(ws, 'C23', cogo_statement)
    safe_set(ws, 'C25', clean_line(idr_info.get('remarks', '')))
    for i in range(PDF_ROW_COUNT):
        excel_row = 13 + i
        row = rows[i] if i < len(rows) else {}
        code = clean_line(row.get('item_code', ''))
        if code == 'Custom / Manual':
            code = ''
        desc = clean_line(row.get('item_description', ''))
        if desc == 'Custom / Manual':
            desc = ''
        qty = clean_line(row.get('quantity', ''))
        unit = normalize_unit(row.get('unit', ''))
        qty_unit = clean_line(f'{qty} {unit}') if qty or unit else ''
        safe_set(ws, f'B{excel_row}', code)
        safe_set(ws, f'C{excel_row}', clean_line(row.get('fund_code', '')))
        safe_set(ws, f'D{excel_row}', desc)
        safe_set(ws, f'F{excel_row}', clean_line(row.get('location', '')))
        safe_set(ws, f'H{excel_row}', qty_unit)
        safe_set(ws, f'J{excel_row}', clean_line(row.get('evidence', '')))
    ws.row_dimensions[12].height = max(ws.row_dimensions[12].height or 0, 25)
    for table_row in range(13, 19):
        ws.row_dimensions[table_row].height = max(ws.row_dimensions[table_row].height or 0, 22)
    ws.row_dimensions[19].height = max(ws.row_dimensions[19].height or 0, 18)
    ws.row_dimensions[20].height = max(ws.row_dimensions[20].height or 0, 18)
    for remarks_row in range(21, 25):
        ws.row_dimensions[remarks_row].height = max(ws.row_dimensions[remarks_row].height or 0, 17)
    for remarks_row in range(23, 26):
        ws.row_dimensions[remarks_row].height = max(ws.row_dimensions[remarks_row].height or 0, 18)
    try:
        format_item_description_cells(ws)
    except Exception:
        pass
    try:
        format_quantity_cells(ws)
    except Exception:
        pass
    try:
        format_evidence_cells(ws)
    except Exception:
        pass
    increase_entire_form_font_size(ws, points=1.5)
    prepare_exact_print_layout(wb, ws)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def find_libreoffice_executable():
    return shutil.which('libreoffice') or shutil.which('soffice')

def convert_xlsx_bytes_to_pdf(xlsx_bytes):
    executable = find_libreoffice_executable()
    if not executable:
        raise RuntimeError('LibreOffice is required for exact Excel-to-PDF output. Add an apt packages file with libreoffice installed.')
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        xlsx_path = tmpdir_path / 'filled_idr.xlsx'
        pdf_path = tmpdir_path / 'filled_idr.pdf'
        xlsx_path.write_bytes(xlsx_bytes.getvalue())
        result = subprocess.run([executable, '--headless', '--convert-to', 'pdf', '--outdir', str(tmpdir_path), str(xlsx_path)], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=90)
        if result.returncode != 0 or not pdf_path.exists():
            raise RuntimeError(f'LibreOffice could not convert the filled template to PDF.\n\nstdout: {result.stdout}\n\nstderr: {result.stderr}')
        return io.BytesIO(pdf_path.read_bytes())

def _draw_pdf_text(c, value, x, y, max_width, size=8, bold=False):
    """Draw one line, shrinking only enough to remain inside its field."""
    value = clean_line(value)
    if not value:
        return

    font_name = 'Helvetica-Bold' if bold else 'Helvetica'
    draw_size = float(size)

    while draw_size > 5 and stringWidth(value, font_name, draw_size) > max_width:
        draw_size -= 0.25

    c.setFillColor(black)
    c.setFont(font_name, draw_size)
    c.drawString(x, y, value)


def _draw_wrapped_pdf_text(c, value, x, y, max_width, max_height, size=7, leading=8):
    """Wrap text inside a fixed rectangle without crossing its borders."""
    value = str(value or '').strip()
    if not value:
        return

    lines = []
    for paragraph in value.splitlines():
        words = paragraph.split()
        if not words:
            lines.append('')
            continue

        current = ''
        for word in words:
            candidate = f'{current} {word}'.strip()
            if stringWidth(candidate, 'Helvetica', size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = word
        if current:
            lines.append(current)

    max_lines = max(1, int(max_height // leading))
    lines = lines[:max_lines]

    c.setFillColor(black)
    c.setFont('Helvetica', size)

    current_y = y
    for line in lines:
        c.drawString(x, current_y, line)
        current_y -= leading


def _contract_suffix(value):
    value = normalize_contract_input(value)
    if not value:
        return ''
    return value.split('-')[-1][-5:]


def _build_flat_bc628_overlay(metadata, idr_info, rows):
    """
    Build an overlay measured directly against the supplied 792 x 612
    bc-628-flat.pdf. These coordinates were verified by rendering the merged
    result back to an image.
    """
    page_width = 792.0
    page_height = 612.0

    overlay = io.BytesIO()
    c = canvas.Canvas(overlay, pagesize=(page_width, page_height))

    report_date = format_report_date(idr_info.get('date', ''))
    inspected_by = clean_line(idr_info.get('inspected_by', ''))
    measured_by = clean_line(idr_info.get('measured_by', ''))
    calculated_by = clean_line(idr_info.get('calculated_by', ''))

    # Upper-left fields.
    _draw_pdf_text(c, report_date, 31, 508, 80, 8)
    _draw_pdf_text(c, clean_line(idr_info.get('contractor', '')), 31, 475, 250, 8)
    _draw_pdf_text(c, idr_info.get('weather', ''), 31, 442, 250, 8)

    # Initials and dates.
    _draw_pdf_text(c, inspected_by, 327, 499, 65, 8)
    _draw_pdf_text(c, report_date if inspected_by else '', 399, 499, 65, 8)

    _draw_pdf_text(c, measured_by, 327, 484, 65, 8)
    _draw_pdf_text(c, report_date if measured_by else '', 399, 484, 65, 8)

    _draw_pdf_text(c, calculated_by, 327, 467.5, 65, 8)
    _draw_pdf_text(c, report_date if calculated_by else '', 399, 467.5, 65, 8)

    # Upper-right contract information.
    _draw_pdf_text(c, metadata.get('county', ''), 550, 508, 100, 8)
    _draw_pdf_text(c, metadata.get('key_route', ''), 659, 508, 100, 8)
    _draw_pdf_text(c, metadata.get('marked_route', ''), 550, 475, 205, 8)
    _draw_pdf_text(c, metadata.get('district', ''), 550, 442, 28, 8)
    _draw_pdf_text(
        c,
        _contract_suffix(metadata.get('item_contract', '')),
        592,
        442,
        76,
        8,
    )
    _draw_pdf_text(c, metadata.get('state_job', ''), 678, 442, 80, 7.5)
    _draw_pdf_text(c, metadata.get('federal_project', ''), 550, 409, 205, 8)

    # Six table rows.
    row_baselines = [349, 327.5, 306, 284.5, 263, 241]

    for row_index in range(PDF_ROW_COUNT):
        row = rows[row_index] if row_index < len(rows) else {}
        baseline = row_baselines[row_index]

        item_code = clean_line(row.get('item_code', ''))
        if item_code.upper() == 'CUSTOM / MANUAL':
            item_code = ''

        description = clean_line(row.get('item_description', ''))
        if description.upper() == 'CUSTOM / MANUAL':
            description = ''

        quantity = clean_line(row.get('quantity', ''))
        unit = normalize_unit(row.get('unit', ''))
        quantity_and_unit = clean_line(f'{quantity} {unit}') if quantity or unit else ''

        _draw_pdf_text(c, item_code, 31, baseline, 80, 7)
        _draw_pdf_text(c, row.get('fund_code', ''), 117, baseline, 45, 7)

        _draw_wrapped_pdf_text(
            c,
            description,
            168,
            baseline + 4,
            116,
            17,
            size=6.8,
            leading=7,
        )
        _draw_wrapped_pdf_text(
            c,
            row.get('location', ''),
            291,
            baseline + 4,
            114,
            17,
            size=6.8,
            leading=7,
        )
        _draw_pdf_text(c, quantity_and_unit, 413, baseline, 79, 7)
        evidence_text = clean_line(row.get('evidence', ''))
        if evidence_text.lower() in {'none', 'n/a', 'na'}:
            evidence_text = ''

        _draw_wrapped_pdf_text(
            c,
            evidence_text,
            499,
            baseline + 5,
            208,
            17,
            size=8.0,
            leading=8.5,
        )

    # Estimated/final measurement selection.
    measurement_type = clean_line(idr_info.get('measurement_type', ''))
    selected_numbers = ', '.join(selected_item_codes(rows))

    if measurement_type == 'Estimated progress measurement':
        c.setFont('Helvetica-Bold', 10)
        c.drawString(61, 224, 'X')
        _draw_pdf_text(c, selected_numbers, 269, 222, 438, 7)

    elif measurement_type == 'Final field measurement':
        c.setFont('Helvetica-Bold', 10)
        c.drawString(61, 204, 'X')
        _draw_pdf_text(c, selected_numbers, 226, 202, 480, 7)

    # Remarks and optional COGO statement, entirely inside the remarks box.
    remarks_parts = []

    if clean_line(idr_info.get('cogo_statement_option', '')) == 'Yes':
        cogo_year = clean_line(idr_info.get('cogo_version_year', ''))
        version_text = f' {cogo_year}' if cogo_year else ''
        remarks_parts.append(
            'Used Cogo Area Calc Tool Trimble Access Version'
            f'{version_text} Area Calculated from a list of points shot around the perimeter '
            '(attached Area Calculation, pointlist, and coordinates measured quantity '
            'compares to plan quantity.)'
        )

    typed_remarks = clean_line(idr_info.get('remarks', ''))
    if typed_remarks:
        remarks_parts.append(typed_remarks)

    remarks_text = '\n\n'.join(remarks_parts)

    _draw_wrapped_pdf_text(
        c,
        remarks_text,
        31,
        168,
        730,
        102,
        size=8.7,
        leading=10.0,
    )

    # Replace the static printed date from the flattened template.
    c.setFillColor(white)
    c.rect(26, 22, 115, 16, fill=1, stroke=0)
    c.setFillColor(black)
    _draw_pdf_text(c, f'Printed {report_date}', 28, 27, 110, 6.5)

    c.save()
    overlay.seek(0)
    return overlay


def make_exact_idr_pdf(metadata, idr_info, rows):
    """
    Fill the exact browser-viewable bc-628-flat.pdf.

    The output is a normal PDF, not XFA, so it opens directly in Chrome, Edge,
    Firefox, Streamlit, phones, and Adobe Reader.
    """
    if not FLAT_PDF_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            'Missing bc-628-flat.pdf. Put the exact flattened BC-628 PDF in '
            'the same folder as app.py.'
        )

    template_reader = PdfReader(str(FLAT_PDF_TEMPLATE_PATH))
    if not template_reader.pages:
        raise RuntimeError('bc-628-flat.pdf does not contain a page.')

    template_page = template_reader.pages[0]

    page_width = float(template_page.mediabox.width)
    page_height = float(template_page.mediabox.height)

    if round(page_width, 1) != 792.0 or round(page_height, 1) != 612.0:
        raise RuntimeError(
            f'Unexpected BC-628 page size: {page_width} x {page_height}. '
            'Use the exact bc-628-flat.pdf supplied with this app.'
        )

    overlay_reader = PdfReader(_build_flat_bc628_overlay(metadata, idr_info, rows))
    template_page.merge_page(overlay_reader.pages[0])

    writer = PdfWriter()
    writer.add_page(template_page)

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def make_pay_items_excel(metadata, pay_items):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        pay_items.to_excel(writer, index=False, sheet_name='Pay Items')
        info_df = pd.DataFrame([{'field': k, 'value': v} for k, v in metadata.items()])
        info_df.to_excel(writer, index=False, sheet_name='Job Info')
        workbook = writer.book
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes(1, 0)
            ws.autofilter(0, 0, 1000, 20)
            header_format = workbook.add_format({'bold': True, 'border': 1, 'text_wrap': True})
            cols = pay_items.columns if sheet_name == 'Pay Items' else info_df.columns
            for col_num, col in enumerate(cols):
                ws.write(0, col_num, col, header_format)
                ws.set_column(col_num, col_num, 24)
            if sheet_name == 'Pay Items':
                ws.set_column(2, 2, 48)
    output.seek(0)
    return output
st.set_page_config(page_title='IDOT Job IDR Generator', page_icon='📄', layout='wide')
st.title('IDOT Job IDR Generator')
st.write('Enter an IDOT job/contract number or paste the direct IDOT contract URL. The website fills the exact browser-viewable BC-628 flat PDF while preserving the existing job lookup, pay-item, evidence, custom-entry, quantity-check, COGO, remarks, and Excel-backup functions.')
with st.sidebar:
    st.header('Job Lookup')
    job_number = st.text_input(
        'IDOT Job / Contract Number or Contract Detail URL',
        placeholder='Example: 62K33, 001-62K33, or paste the IDOT contract URL',
    )

    if st.button('Build / Refresh Contract Index', use_container_width=True):
        try:
            with st.spinner('Building the full IDOT contract index. This may take a few minutes...'):
                session = make_session()

                # Refresh the letting list instead of relying on an old Streamlit cache.
                letting_links = get_all_archive_letting_links_newest_first(session)

                if not letting_links:
                    raise RuntimeError('IDOT did not return any letting/archive pages.')

                result = build_sqlite_contract_index_for_lettings(
                    letting_links=letting_links,
                    max_pages_per_letting=FULL_INDEX_MAX_PAGES_PER_LETTING,
                    meta_key='full_contract_index',
                    meta_ttl_seconds=FULL_INDEX_TTL_SECONDS,
                    force=True,
                    source_suffix='manual-full-index',
                )

                # A job that previously failed may be stored in contract_misses.
                # Clear those misses so the newly built index is checked immediately.
                sqlite_clear_contract_misses()
                st.cache_data.clear()

            st.success(
                f"Index complete: checked {result.get('checked_lettings', 0)} letting(s), "
                f"{result.get('checked_pages', 0)} page(s), and saved "
                f"{result.get('saved_contracts', 0)} contract entries."
            )
        except Exception as e:
            st.error(f'Index build failed: {e}')
if 'metadata' not in st.session_state:
    st.session_state.metadata = None
if 'pay_items' not in st.session_state:
    st.session_state.pay_items = pd.DataFrame()
if 'match' not in st.session_state:
    st.session_state.match = None
if st.button('Find IDOT Job'):
    try:
        with st.spinner('Fast lookup: checking local index and recent IDOT letting pages...'):
            metadata, pay_items, match = fetch_idot_job(job_number)
        st.session_state.metadata = metadata
        st.session_state.pay_items = pay_items
        st.session_state.match = match
        clear_idr_row_state()
        st.success(f"Found {metadata.get('item_contract', job_number)} with {len(pay_items)} pay items.")
        if match.get('letting_url'):
            st.info(f"Found on letting/archive page: {match.get('letting', '')}, page {match.get('page', '')}")
        else:
            st.info(f"Found using: {match.get('letting', '')}")
    except Exception as e:
        st.error(str(e))
metadata = st.session_state.metadata
pay_items = st.session_state.pay_items
if metadata is not None:
    st.subheader('Job Information')
    required_fields = {'County': metadata.get('county', ''), 'Section': metadata.get('key_route', ''), 'Route': metadata.get('marked_route', ''), 'District': metadata.get('district', ''), 'Contract No.': metadata.get('item_contract', ''), 'Job No.': metadata.get('state_job', ''), 'Project': metadata.get('federal_project', '')}
    missing_fields = [name for name, value in required_fields.items() if not value]
    if missing_fields:
        st.warning('Some job fields did not parse correctly: ' + ', '.join(missing_fields))
        with st.expander('Parser debug info'):
            st.write(metadata)
    col1, col2, col3 = st.columns(3)
    with col1:
        st.write(f"**County:** {metadata.get('county', '')}")
        st.write(f"**Section:** {metadata.get('key_route', '')}")
        st.write(f"**Route:** {metadata.get('marked_route', '')}")
    with col2:
        st.write(f"**District:** {metadata.get('district', '')}")
        st.write(f"**Contract No.:** {metadata.get('item_contract', '')}")
        st.write(f"**Job No.:** {metadata.get('state_job', '')}")
    with col3:
        st.write(f"**Project:** {metadata.get('federal_project', '')}")
        st.write(f"**Working Days:** {metadata.get('working_days', '')}")
if metadata is not None and (not pay_items.empty):
    idr_info = build_idr_header_form()
    st.divider()
    rows = build_idr_rows_form(pay_items)
    st.divider()
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        pay_items_file = make_pay_items_excel(metadata, pay_items)
        st.download_button(label='Download IDOT Pay Item Table', data=pay_items_file, file_name=f"{metadata.get('item_contract', 'idot_job')}_pay_items.xlsx", mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    with col_b:
        try:
            filled_xlsx = fill_exact_idr_workbook(metadata, idr_info, rows)
            st.download_button(label='Download Filled IDR Excel Backup', data=filled_xlsx, file_name=format_xlsx_filename(metadata.get('item_contract', 'IDOT')), mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except Exception as e:
            st.error(f'Could not prepare filled Excel backup: {e}')
    with col_c:
        try:
            pdf_file = make_exact_idr_pdf(metadata, idr_info, rows)
            st.download_button(label='Download Filled BC-628 PDF', data=pdf_file, file_name=format_pdf_filename(metadata.get('item_contract', 'IDOT')), mime='application/pdf')
        except Exception as e:
            st.error(f'Could not generate BC-628 PDF: {e}')